from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .models import TakeoffAsset, TakeoffDocument


HEADER_FILL = PatternFill("solid", fgColor="6FA8DC")
QA_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(name="Carlito", size=11, bold=True, color="000000")
BODY_FONT = Font(name="Carlito", size=11, color="000000")
MAX_COLUMN_WIDTH = 72
PDF_POINTS_TO_90_DPI = 90 / 72

TABLE_NAMES = {
    "Resumen Takeoff": "ResumenTakeoffTable",
    "Codigos Elementos": "CodigosElementosTable",
    "Registro Activos": "RegistroActivosTable",
    "Hojas de Dibujo": "HojasDibujoTable",
    "Inventario Edificios": "InventarioEdificiosTable",
    "Registros Edificios": "RegistrosEdificiosTable",
    "Por Tipo Edificio": "PorTipoEdificioTable",
    "Por Piso": "PorPisoTable",
    "Por Area": "PorAreaTable",
    "Analisis Cantidades": "AnalisisCantidadesTable",
    "Preguntas": "PreguntasTable",
    "Panel": "PanelTable",
    "Breaker": "BreakerTable",
    "Analisis Unifilar": "AnalisisUnifilarTable",
    "Distancias Cableado": "DistanciasCableadoTable",
    "Cronograma Equipos": "CronogramaEquiposTable",
    "Cadena Unifilar": "CadenaUnifilarTable",
    "Estimados Cableado": "EstimadosCableadoTable",
    "Estado Entregables": "EstadoEntregablesTable",
}

SHEET_COLUMN_WIDTHS = {
    "Resumen Takeoff": (9, 18, 18, 12, 44, 12, 10, 17, 9, 16, 17, 15, 16, 16, 16, 18, 16, 14, 16, 18, 15, 14, 25, 54),
    "Codigos Elementos": (40, 10, 15, 44, 12, 16, 24, 24, 64, 10, 42, 26, 22),
    "Registro Activos": (40, 22, 10, 24, 34, 42, 24, 24, 64, 10, 12, 12, 12, 12, 22, 24, 22, 16, 18, 15, 44, 40, 20, 28, 18, 28, 58),
    "Hojas de Dibujo": (10, 25, 58, 28, 18, 22, 12, 48),
    "Inventario Edificios": (28, 20, 22, 22, 30, 22),
    "Registros Edificios": (30, 26, 45, 42),
    "Por Tipo Edificio": (28, 25, 44, 14),
    "Por Piso": (20, 25, 44, 14),
    "Por Area": (17, 48, 25, 44, 14),
    "Analisis Cantidades": (17, 48, 14, 22, 25, 18, 44, 14, 32),
    "Preguntas": (22, 72, 14, 35, 52),
    "Panel": (38, 72),
    "Breaker": (26, 28, 25, 58),
    "Analisis Unifilar": (28, 20, 22, 22, 42, 18, 72),
    "Distancias Cableado": (20, 22, 22, 14, 12, 22, 34, 22, 58),
    "Cronograma Equipos": (30, 25, 42, 24, 20, 28, 16, 25, 28, 58),
    "Cadena Unifilar": (20, 22, 22, 22, 28, 22, 22, 22, 58),
    "Estimados Cableado": (22, 24, 22, 22, 20, 12, 24, 34, 22, 58),
    "Estado Entregables": (32, 20, 30, 58),
}

SUMMARY_HEADERS = (
    "Item #",
    "Ref. Sheet",
    "Detail Sheet Ref.",
    "CSI Ref.",
    "Description",
    "Quantity",
    "Wastage",
    "Quantity w/ Wastage",
    "Unit",
    "Unit Material Cost",
    "Total Material Cost",
    "Unit Labor Cost",
    "Total Labor Cost",
    "Total Item Cost",
    "Division Sub Total",
    "Supplier",
    "Supplier Country",
    "Source Currency",
    "Source Unit Price",
    "Source Basis",
    "FX Rate to USD",
    "USD Unit Price",
    "Supplier Link",
    "Sourcing Notes",
)

AUDIT_HEADERS = (
    "unit_id",
    "legend_entry_id",
    "measurement_kind",
    "code",
    "description",
    "page",
    "sheet",
    "area_code",
    "area",
    "level",
    "method",
    "confidence",
    "quantity",
    "unit",
    "path_length_pdf_points",
    "scale_kind",
    "scale_source_page",
    "scale_source_sheet",
    "scale_source_text",
    "scale_real_units_per_pdf_point",
)

EQUIPMENT_KEYWORDS = (
    "PANEL",
    "TABLERO",
    "TRANSFORM",
    "BREAKER",
    "SWITCHBOARD",
    "DISTRIB",
    "SECCIONADORA",
    "GENERATOR",
    "GENERADOR",
    "UPS",
    "ATS",
)


def _static_number(value: float) -> int | float:
    rounded = round(value)
    if abs(value - rounded) <= 1e-9:
        return int(rounded)
    return round(value, 6)


def _bounded_text(value: Any, limit: int = 2_000) -> Any:
    if not isinstance(value, str) or len(value) <= limit:
        return value
    return value[: limit - 1] + "…"


def _coordinate_text(asset: TakeoffAsset) -> str:
    if asset.bbox is not None:
        return (
            f"x0={asset.bbox.x0 * PDF_POINTS_TO_90_DPI:.1f}; "
            f"y0={asset.bbox.y0 * PDF_POINTS_TO_90_DPI:.1f}; "
            f"x1={asset.bbox.x1 * PDF_POINTS_TO_90_DPI:.1f}; "
            f"y1={asset.bbox.y1 * PDF_POINTS_TO_90_DPI:.1f}"
        )
    if asset.x is not None and asset.y is not None:
        return (
            f"x={asset.x * PDF_POINTS_TO_90_DPI:.1f}; "
            f"y={asset.y * PDF_POINTS_TO_90_DPI:.1f}"
        )
    points = asset.path or []
    shown = points[:20]
    text = " -> ".join(
        f"({point.x * PDF_POINTS_TO_90_DPI:.1f},"
        f"{point.y * PDF_POINTS_TO_90_DPI:.1f})"
        for point in shown
    )
    if len(points) > len(shown):
        text += f" -> … ({len(points)} puntos)"
    return text


def _asset_dimensions(asset: TakeoffAsset) -> tuple[float, float, float | None, float | None]:
    center = asset.center()
    if asset.bbox is None:
        return (
            center.x * PDF_POINTS_TO_90_DPI,
            center.y * PDF_POINTS_TO_90_DPI,
            None,
            None,
        )
    return (
        center.x * PDF_POINTS_TO_90_DPI,
        center.y * PDF_POINTS_TO_90_DPI,
        (asset.bbox.x1 - asset.bbox.x0) * PDF_POINTS_TO_90_DPI,
        (asset.bbox.y1 - asset.bbox.y0) * PDF_POINTS_TO_90_DPI,
    )


def _source_name(takeoff: TakeoffDocument) -> str:
    return takeoff.source.pdf or "Planos cargados"


def _summary_groups(
    assets: Iterable[TakeoffAsset],
) -> list[tuple[tuple[str, str, str, str, str], float, set[str]]]:
    quantities: dict[tuple[str, str, str, str, str], float] = defaultdict(float)
    sheets: dict[tuple[str, str, str, str, str], set[str]] = defaultdict(set)
    for asset in assets:
        key = (
            asset.legend_entry_id,
            asset.code,
            asset.description,
            asset.measurement_kind,
            asset.unit,
        )
        quantities[key] += asset.quantity
        sheets[key].add(asset.sheet)
    return [
        (key, quantities[key], sheets[key])
        for key in sorted(quantities, key=lambda item: (item[1], item[0]))
    ]


def _aggregate(
    assets: Iterable[TakeoffAsset],
    key_fields: Sequence[str],
) -> list[tuple[tuple[str, ...], float]]:
    totals: dict[tuple[str, ...], float] = defaultdict(float)
    for asset in assets:
        key = tuple(str(getattr(asset, field)) for field in key_fields)
        totals[key] += asset.quantity
    return sorted(totals.items(), key=lambda item: item[0])


def _is_equipment(asset: TakeoffAsset) -> bool:
    haystack = f"{asset.code} {asset.description}".upper()
    return any(keyword in haystack for keyword in EQUIPMENT_KEYWORDS)


def _style_cell(
    cell: Cell,
    *,
    header: bool,
    body_fill: PatternFill | None = None,
    body_wrap: bool = False,
) -> None:
    cell.font = HEADER_FONT if header else BODY_FONT
    if header:
        cell.fill = HEADER_FILL
    elif body_fill is not None:
        cell.fill = body_fill
    cell.alignment = Alignment(
        horizontal="center" if header else "left",
        vertical="center",
        wrap_text=header or body_wrap,
    )


def _table_name(index: int, title: str) -> str:
    if title in TABLE_NAMES:
        return TABLE_NAMES[title]
    safe = "".join(character for character in title if character.isalnum())
    return f"CuadraBot{index:02d}{safe[:20]}"


def _write_sheet(
    workbook: Workbook,
    *,
    index: int,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    placeholder: Sequence[Any] | None = None,
    body_fill: PatternFill | None = None,
    body_wrap: bool = False,
) -> Worksheet:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(headers))
    materialized = list(rows)
    if not materialized and placeholder is not None:
        materialized.append(list(placeholder))
    for row in materialized:
        worksheet.append([_bounded_text(value) for value in row])

    worksheet.freeze_panes = None
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 32
    for cell in worksheet[1]:
        _style_cell(cell, header=True)
    for row_number in range(2, worksheet.max_row + 1):
        for cell in worksheet[row_number]:
            _style_cell(
                cell,
                header=False,
                body_fill=body_fill,
                body_wrap=body_wrap,
            )
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.alignment = Alignment(horizontal="right", vertical="center")

    exact_widths = SHEET_COLUMN_WIDTHS.get(title)
    for column_number, header in enumerate(headers, start=1):
        sample_values = [header]
        for row_number in range(2, min(worksheet.max_row, 250) + 1):
            value = worksheet.cell(row=row_number, column=column_number).value
            if value is not None:
                sample_values.append(str(value))
        width = (
            exact_widths[column_number - 1]
            if exact_widths is not None
            else min(
                MAX_COLUMN_WIDTH,
                max(
                    10,
                    max((len(str(value)) for value in sample_values), default=10) + 2,
                ),
            )
        )
        worksheet.column_dimensions[worksheet.cell(1, column_number).column_letter].width = width

    if worksheet.max_row >= 2:
        table = Table(
            displayName=_table_name(index, title),
            ref=f"A1:{worksheet.cell(worksheet.max_row, len(headers)).coordinate}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(headers)).coordinate}"
    return worksheet


def _audit_rows(takeoff: TakeoffDocument) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for asset in takeoff.assets:
        evidence = asset.scale_evidence
        rows.append(
            [
                asset.unit_id,
                asset.legend_entry_id,
                asset.measurement_kind,
                asset.code,
                asset.description,
                asset.page,
                asset.sheet,
                asset.area_code,
                asset.area,
                asset.level,
                asset.method,
                asset.confidence,
                _static_number(asset.quantity),
                asset.unit,
                _static_number(asset.display_path_length_points())
                if asset.measurement_kind == "linear"
                else None,
                evidence.kind if evidence is not None else None,
                evidence.page if evidence is not None else None,
                evidence.sheet if evidence is not None else None,
                evidence.source_text if evidence is not None else None,
                _static_number(evidence.derived_real_units_per_pdf_point())
                if evidence is not None
                else None,
            ]
        )
    return rows


def build_takeoff_workbook(takeoff: TakeoffDocument, output_path: Path) -> None:
    """Write the static ORTEGA-format customer workbook plus a hidden audit tab."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"

    assets = list(takeoff.assets)
    groups = _summary_groups(assets)
    summary_rows: list[list[Any]] = []
    for item_number, (key, quantity, sheets) in enumerate(groups, start=1):
        _legend_id, code, description, _measurement_kind, unit = key
        ref_sheet = next(iter(sheets)) if len(sheets) == 1 else "Múltiples hojas"
        static_quantity = _static_number(quantity)
        summary_rows.append(
            [
                item_number,
                ref_sheet,
                key[0],
                None,
                f"{code} — {description}",
                static_quantity,
                0,
                static_quantity,
                unit,
                0,
                0,
                0,
                0,
                0,
                0,
                None,
                None,
                "USD",
                0,
                "Conteo de planos" if unit == "EA" else "Medición de planos",
                1,
                0,
                None,
                (
                    "Precios y mano de obra fuera del alcance; cantidad basada "
                    "en los planos analizados."
                ),
            ]
        )
    _write_sheet(
        workbook,
        index=1,
        title="Resumen Takeoff",
        headers=SUMMARY_HEADERS,
        rows=summary_rows,
        placeholder=[None, None, None, None, "Sin cantidades incluidas"]
        + [0] * 19,
    )

    code_rows = []
    asset_rows = []
    for number, asset in enumerate(assets, start=1):
        coordinate_text = _coordinate_text(asset)
        x_value, y_value, width, height = _asset_dimensions(asset)
        code_rows.append(
            [
                asset.unit_id,
                number,
                asset.area_code,
                asset.area,
                "Conteo" if asset.measurement_kind == "count" else "Lineal",
                asset.level,
                asset.sheet,
                asset.sheet,
                asset.sheet,
                _static_number(asset.quantity),
                asset.description,
                asset.code,
                coordinate_text,
            ]
        )
        asset_rows.append(
            [
                asset.unit_id,
                asset.visible_label,
                number,
                asset.code,
                asset.description,
                asset.description,
                asset.sheet,
                asset.sheet,
                asset.sheet,
                asset.page,
                _static_number(x_value),
                _static_number(y_value),
                _static_number(width) if width is not None else None,
                _static_number(height) if height is not None else None,
                coordinate_text,
                None,
                None,
                asset.level,
                None,
                asset.area_code,
                asset.area,
                "Asignación del análisis de planos",
                None,
                f"{asset.area_code}-{asset.code}",
                asset.method,
                asset.confidence,
                asset.notes,
            ]
        )

    _write_sheet(
        workbook,
        index=2,
        title="Codigos Elementos",
        headers=(
            "ID",
            "Numero",
            "Codigo de Area",
            "Area",
            "Tipo",
            "Piso",
            "Hoja",
            "Hoja de Piso",
            "Nombre de Plano",
            "Cantidad",
            "Descripcion",
            "Elemento",
            "Coordenadas",
        ),
        rows=code_rows,
        placeholder=["Sin elementos", None, None, None, None, None, None, None, None, 0],
    )
    _write_sheet(
        workbook,
        index=3,
        title="Registro Activos",
        headers=(
            "ID Activo",
            "Etiqueta Visible",
            "Numero",
            "Codigo",
            "Elemento",
            "Descripcion",
            "Hoja",
            "Hoja de Piso",
            "Nombre de Plano",
            "Pagina",
            "X90",
            "Y90",
            "Ancho 90",
            "Alto 90",
            "Coordenadas",
            "Edificio",
            "Tipo de Edificio",
            "Piso",
            "Tipo de Apartamento",
            "Codigo de Area",
            "Area",
            "Base de Area",
            "Distancia Etiqueta Area",
            "Grupo de Conteo",
            "Metodo Fuente",
            "Estado QA",
            "Notas",
        ),
        rows=asset_rows,
        placeholder=["Sin activos"] + [None] * 26,
    )

    assets_by_page: dict[int, list[TakeoffAsset]] = defaultdict(list)
    legend_sheets_by_page: dict[int, set[str]] = defaultdict(set)
    for asset in assets:
        assets_by_page[asset.page].append(asset)
    for entry in takeoff.legend_entries:
        legend_sheets_by_page[entry.page].add(entry.sheet)
    drawing_rows = []
    for page in range(1, takeoff.source.page_count + 1):
        page_assets = assets_by_page.get(page, [])
        sheets = {asset.sheet for asset in page_assets} | legend_sheets_by_page.get(page, set())
        kinds = {asset.measurement_kind for asset in page_assets}
        classification = (
            "conteo y medición"
            if len(kinds) > 1
            else "conteo"
            if kinds == {"count"}
            else "medición"
            if kinds == {"linear"}
            else "referencia/no clasificada"
        )
        levels = sorted({asset.level for asset in page_assets})
        drawing_rows.append(
            [
                page,
                ", ".join(sorted(sheets)) or f"PDF-{page}",
                ", ".join(sorted(sheets)) or "Título no disponible en takeoff.json",
                classification,
                ", ".join(levels),
                None,
                "Sí" if page_assets else "Por verificar",
                (
                    f"{len(page_assets)} colocaciones o recorridos incluidos."
                    if page_assets
                    else "Sin colocaciones incluidas; revisar clasificación en el PDF."
                ),
            ]
        )
    _write_sheet(
        workbook,
        index=4,
        title="Hojas de Dibujo",
        headers=(
            "Pagina",
            "Numero de Hoja",
            "Nombre de Plano",
            "Clasificacion",
            "Piso",
            "Tipo de Apartamento",
            "Contable",
            "Notas",
        ),
        rows=drawing_rows,
    )

    all_pages = f"1-{takeoff.source.page_count}"
    _write_sheet(
        workbook,
        index=5,
        title="Inventario Edificios",
        headers=(
            "Tipo de Edificio",
            "Cantidad Observada",
            "Cantidad Rango Inferido",
            "Cantidad Seleccionada",
            "IDs Faltantes o Sospechosos",
            "Paginas Fuente",
        ),
        rows=[["No especificado", None, None, None, "No disponible en takeoff.json", all_pages]],
    )
    _write_sheet(
        workbook,
        index=6,
        title="Registros Edificios",
        headers=("Nombre de Edificio", "Tipo de Edificio", "Estado Fuente", "Base"),
        rows=[
            [
                "No especificado",
                "No especificado",
                "Requiere revisión del cajetín",
                "No inferido por el contrato de conteo",
            ]
        ],
    )

    type_rows = [
        [
            "No especificado",
            key[1],
            f"{key[2]} [{key[4]}]",
            _static_number(quantity),
        ]
        for key, quantity, _sheets in groups
    ]
    _write_sheet(
        workbook,
        index=7,
        title="Por Tipo Edificio",
        headers=("Tipo", "Codigo", "Descripcion", "Cantidad"),
        rows=type_rows,
        placeholder=["No especificado", None, "Sin cantidades", 0],
    )

    floor_rows = [
        [level, code, f"{description} [{unit}]", _static_number(quantity)]
        for (level, code, description, _kind, unit), quantity in _aggregate(
            assets, ("level", "code", "description", "measurement_kind", "unit")
        )
    ]
    _write_sheet(
        workbook,
        index=8,
        title="Por Piso",
        headers=("Piso", "Codigo", "Descripcion", "Cantidad"),
        rows=floor_rows,
        placeholder=[None, None, "Sin cantidades", 0],
    )

    area_rows = [
        [
            area_code,
            area,
            code,
            f"{description} [{unit}]",
            _static_number(quantity),
        ]
        for (area_code, area, code, description, _kind, unit), quantity in _aggregate(
            assets,
            ("area_code", "area", "code", "description", "measurement_kind", "unit"),
        )
    ]
    _write_sheet(
        workbook,
        index=9,
        title="Por Area",
        headers=("Codigo de Area", "Area", "Codigo", "Descripcion", "Cantidad"),
        rows=area_rows,
        placeholder=[None, None, None, "Sin cantidades", 0],
    )

    analysis_rows = []
    analysis_groups: dict[tuple[str, str, str, str, str, str], list[TakeoffAsset]] = defaultdict(list)
    for asset in assets:
        analysis_groups[
            (
                asset.area_code,
                asset.area,
                asset.measurement_kind,
                asset.code,
                asset.description,
                asset.unit,
            )
        ].append(asset)
    for key in sorted(analysis_groups):
        area_code, area, kind, code, description, unit = key
        grouped_assets = analysis_groups[key]
        quantity = sum(asset.quantity for asset in grouped_assets)
        sources = sorted({f"{asset.sheet}, PDF p.{asset.page}" for asset in grouped_assets})
        labels = sorted({asset.visible_label for asset in grouped_assets if asset.visible_label})
        analysis_rows.append(
            [
                area_code,
                area,
                f"{'Conteo' if kind == 'count' else 'Lineal'} ({unit})",
                None,
                code,
                ", ".join(labels[:5]),
                description,
                _static_number(quantity),
                "; ".join(sources[:10]),
            ]
        )
    _write_sheet(
        workbook,
        index=10,
        title="Analisis Cantidades",
        headers=(
            "Codigo de Area",
            "Area",
            "Tipo",
            "Tipo de Apartamento",
            "Codigo",
            "Prefijo Visible",
            "Descripcion",
            "Cantidad",
            "Base",
        ),
        rows=analysis_rows,
        placeholder=[None, None, None, None, None, None, "Sin cantidades", 0, None],
    )

    qa_rows: list[list[Any]] = []
    qa_number = 1
    if takeoff.unresolved_symbols:
        qa_rows.append(
            [
                f"Control QA {qa_number}",
                f"{len(takeoff.unresolved_symbols)} símbolos sin resolución fueron excluidos.",
                None,
                "takeoff.json.unresolved_symbols",
                "Revisar antes de compra o construcción.",
            ]
        )
        qa_number += 1
    for limitation in takeoff.limitations:
        qa_rows.append(
            [
                f"Control QA {qa_number}",
                limitation,
                None,
                "takeoff.json.limitations",
                "Revisión requerida antes de compra o construcción.",
            ]
        )
        qa_number += 1
    if not qa_rows:
        qa_rows.append(
            [
                "Control QA 1",
                "No se registraron limitaciones adicionales en takeoff.json.",
                None,
                "Validación determinística de CuadraBot",
                "El conteo sigue requiriendo revisión profesional antes de construcción.",
            ]
        )
    _write_sheet(
        workbook,
        index=11,
        title="Preguntas",
        headers=("Pregunta", "Respuesta", "Unidad", "Base", "Nota QA"),
        rows=qa_rows,
        body_fill=QA_FILL,
        body_wrap=True,
    )

    countable_pages = len(assets_by_page)
    quantity_totals = _aggregate(assets, ("measurement_kind", "unit"))
    panel_rows: list[list[Any]] = [
        ["Proyecto", _source_name(takeoff)],
        ["Fuente", _source_name(takeoff)],
        ["Estado", "Resultado de análisis; requiere revisión profesional"],
        ["Páginas indexadas", takeoff.source.page_count],
        ["Hojas con cantidades", countable_pages],
        ["Hojas sin cantidades", takeoff.source.page_count - countable_pages],
        ["Unidades/recorridos", len(assets)],
        ["Grupos código/unidad", len(groups)],
        ["Método", "Conteo/medición con geometría fuente"],
        ["QA", f"{len(takeoff.unresolved_symbols)} símbolos sin resolver"],
        ["SHA-256 fuente", takeoff.source.sha256],
    ]
    panel_rows.extend(
        [
            f"{'Conteo total' if kind == 'count' else 'Longitud total'} ({unit})",
            _static_number(quantity),
        ]
        for (kind, unit), quantity in quantity_totals
    )
    panel_rows.extend(
        [f"{key[1]} ({key[4]})", _static_number(quantity)]
        for key, quantity, _sheets in groups
    )
    _write_sheet(
        workbook,
        index=12,
        title="Panel",
        headers=("Indicador", "Valor"),
        rows=panel_rows,
        body_wrap=True,
    )

    equipment_assets = [asset for asset in assets if _is_equipment(asset)]
    breaker_rows = [
        [
            "Equipo de distribución",
            asset.visible_label,
            asset.code,
            f"{asset.description} — {asset.area}",
        ]
        for asset in equipment_assets
    ]
    _write_sheet(
        workbook,
        index=13,
        title="Breaker",
        headers=("Seccion", "Dato", "Panel", "Elemento"),
        rows=breaker_rows,
        placeholder=["No derivado", None, None, "Sin equipos de distribución identificados"],
    )

    equipment_pages = sorted({asset.page for asset in equipment_assets})
    _write_sheet(
        workbook,
        index=14,
        title="Analisis Unifilar",
        headers=(
            "Tipo de Edificio",
            "Cantidad Observada",
            "Cantidad Rango Inferido",
            "Cantidad Seleccionada",
            "IDs Faltantes o Sospechosos",
            "Paginas Fuente",
            "Base",
        ),
        rows=[
            [
                "No especificado",
                len(equipment_assets),
                None,
                None,
                "No evaluado como estudio selectivo de coordinación",
                ", ".join(str(page) for page in equipment_pages),
                (
                    "Inventario de equipos detectados; no sustituye un análisis "
                    "unifilar."
                ),
            ]
        ],
    )

    linear_assets = [asset for asset in assets if asset.measurement_kind == "linear"]
    distance_rows = [
        [
            asset.unit_id,
            None,
            None,
            _static_number(asset.quantity),
            asset.unit,
            "Medido",
            asset.scale_evidence.source_text if asset.scale_evidence else None,
            asset.sheet,
            asset.notes,
        ]
        for asset in linear_assets
    ]
    _write_sheet(
        workbook,
        index=15,
        title="Distancias Cableado",
        headers=(
            "Segmento",
            "Desde Nodo",
            "Hasta Nodo",
            "Distancia",
            "Unidad",
            "Estado",
            "Base",
            "Hoja",
            "Notas",
        ),
        rows=distance_rows,
        placeholder=[
            "No medido",
            None,
            None,
            None,
            "m",
            "Fuera de alcance",
            "Sin recorridos lineales incluidos",
            None,
            "Las longitudes requieren escala y ruta verificadas.",
        ],
    )

    equipment_rows = [
        [
            asset.unit_id,
            asset.code,
            asset.description,
            asset.visible_label,
            _static_number(asset.quantity),
            None,
            asset.level,
            asset.sheet,
            asset.visible_label,
            asset.notes,
        ]
        for asset in equipment_assets
    ]
    _write_sheet(
        workbook,
        index=16,
        title="Cronograma Equipos",
        headers=(
            "ID Equipo",
            "Tipo de Equipo",
            "Descripcion",
            "Capacidad/Rating",
            "Cantidad/Capacidad",
            "Edificio",
            "Piso",
            "Hoja",
            "Texto Fuente",
            "Notas",
        ),
        rows=equipment_rows,
        placeholder=["Sin equipos", None, "Sin equipos identificados", None, 0],
    )
    _write_sheet(
        workbook,
        index=17,
        title="Cadena Unifilar",
        headers=(
            "ID Cadena",
            "Desde Nodo",
            "Hasta Nodo",
            "Tipo de Nodo",
            "Referencia Equipo",
            "Capacidad/Rating",
            "Hoja",
            "Coordenadas",
            "Notas",
        ),
        rows=[
            [
                "No trazada",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                "La cadena unifilar no forma parte del conteo unitario.",
            ]
        ],
    )

    cable_rows = [
        [
            asset.unit_id,
            asset.code,
            None,
            None,
            _static_number(asset.quantity),
            asset.unit,
            asset.description,
            asset.scale_evidence.source_text if asset.scale_evidence else None,
            asset.sheet,
            asset.notes,
        ]
        for asset in linear_assets
    ]
    _write_sheet(
        workbook,
        index=18,
        title="Estimados Cableado",
        headers=(
            "ID Cable",
            "Tipo de Segmento",
            "Desde Nodo",
            "Hasta Nodo",
            "Longitud Estimada",
            "Unidad",
            "Tipo de Cable",
            "Base",
            "Hoja",
            "Notas",
        ),
        rows=cable_rows,
        placeholder=[
            "No estimado",
            None,
            None,
            None,
            None,
            "m",
            None,
            "Sin recorridos lineales incluidos",
            None,
            "No se estimaron rutas ni holguras de cableado.",
        ],
    )

    status_rows = [
        ["Registro de hojas", "Completo", "Hojas de Dibujo", f"{takeoff.source.page_count} páginas"],
        ["Fuentes con cantidades", "Completo", "Hojas de Dibujo", f"{countable_pages} hojas"],
        ["Conteo detallado", "Completo", "Registro Activos", "Una fila por colocación o recorrido"],
        ["Códigos unitarios", "Completo", "Codigos Elementos", "ID estable por colocación"],
        ["Resumen por código", "Completo", "Resumen Takeoff", f"{len(groups)} códigos/definiciones"],
        ["Resumen por área", "Completo", "Por Area", f"{len(area_rows)} filas"],
        ["Resumen por piso", "Completo", "Por Piso", f"{len(floor_rows)} filas"],
        ["Inventario de equipos", "Completo" if equipment_assets else "No incluido", "Cronograma Equipos", f"{len(equipment_assets)} unidades"],
        ["Costos de material", "No incluido", "Resumen Takeoff", "Campos conservados en cero"],
        ["Costos de mano de obra", "No incluido", "Resumen Takeoff", "Campos conservados en cero"],
        ["Longitudes de cable", "Completo" if linear_assets else "No incluido", "Estimados Cableado", f"{len(linear_assets)} recorridos"],
        ["Distancias", "Completo" if linear_assets else "No incluido", "Distancias Cableado", "Escala fuente conservada" if linear_assets else "Requiere trazado/escala"],
        ["Símbolos sin resolver", "Verificar" if takeoff.unresolved_symbols else "Completo", "Preguntas", f"{len(takeoff.unresolved_symbols)} símbolos"],
        ["Uso", "Estimación / revisión", "Preguntas", "No es liberación para construcción"],
    ]
    _write_sheet(
        workbook,
        index=19,
        title="Estado Entregables",
        headers=("Entregable", "Estado", "Archivo", "Notas"),
        rows=status_rows,
    )

    audit = _write_sheet(
        workbook,
        index=20,
        title="Takeoff",
        headers=AUDIT_HEADERS,
        rows=_audit_rows(takeoff),
    )
    audit.sheet_state = "hidden"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
