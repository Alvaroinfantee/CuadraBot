from __future__ import annotations

import json
import math
import re
import stat
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

from .annotations import displayed_size
from .models import BoundingBox, Point, TakeoffAsset, TakeoffDocument


MAX_JSON_BYTES = 50 * 1024**2
MAX_WORKBOOK_BYTES = 100 * 1024**2
MAX_PDF_BYTES = 250 * 1024**2
MAX_ZIP_ENTRIES = 10_000
MAX_ZIP_UNCOMPRESSED_BYTES = 250 * 1024**2
MAX_WORKBOOK_SHEETS = 50
MAX_WORKBOOK_ROWS = 200_000
MAX_WORKBOOK_CELLS = 2_000_000
MAX_CELL_TEXT = 32_767
REQUIRED_TAKEOFF_HEADERS = (
    "unit_id",
    "legend_entry_id",
    "measurement_kind",
    "code",
    "description",
    "page",
    "sheet",
    "area_code",
    "area",
    "level",
    "method",
    "confidence",
    "quantity",
    "unit",
    "path_length_pdf_points",
    "scale_kind",
    "scale_source_page",
    "scale_source_sheet",
    "scale_source_text",
    "scale_real_units_per_pdf_point",
)
SUMMARY_TOTAL_KEYS = (
    "quantity",
    "total_quantity",
    "count",
    "total",
    "counted_units",
)
LINEAR_QUANTITY_REL_TOLERANCE = 1e-4
LINEAR_QUANTITY_ABS_TOLERANCE = 1e-3


class ArtifactValidationError(ValueError):
    pass


def _xml_local_name(name: Any) -> str:
    if not isinstance(name, str):
        return ""
    return name.rsplit("}", 1)[-1].rsplit(":", 1)[-1].casefold()


def _reject_external_relationships(payload: bytes) -> None:
    lowered = payload.lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise ArtifactValidationError(
            "takeoff.xlsx contains an unsafe relationship part"
        )
    try:
        root = ElementTree.fromstring(payload)
    except (ElementTree.ParseError, ValueError) as exc:
        raise ArtifactValidationError(
            "takeoff.xlsx contains a malformed relationship part"
        ) from exc
    for element in root.iter():
        if _xml_local_name(element.tag) != "relationship":
            continue
        for attribute, value in element.attrib.items():
            if (
                _xml_local_name(attribute) == "targetmode"
                and value.strip().casefold() == "external"
            ):
                raise ArtifactValidationError(
                    "takeoff.xlsx contains external or active relationships"
                )


def _reject_defined_names(payload: bytes) -> None:
    lowered = payload.lower()
    if b"<!doctype" in lowered or b"<!entity" in lowered:
        raise ArtifactValidationError(
            "takeoff.xlsx contains an unsafe workbook definition"
        )
    try:
        root = ElementTree.fromstring(payload)
    except (ElementTree.ParseError, ValueError) as exc:
        raise ArtifactValidationError(
            "takeoff.xlsx contains malformed workbook definitions"
        ) from exc
    if any(
        _xml_local_name(element.tag) == "definedname"
        for element in root.iter()
    ):
        raise ArtifactValidationError(
            "takeoff.xlsx contains workbook defined names; defined names are "
            "not permitted"
        )


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ArtifactValidationError(
                f"JSON contains duplicate key {key!r}"
            )
        result[key] = value
    return result


def require_regular_file(
    path: Path,
    *,
    allowed_parent: Path,
    max_bytes: int,
    magic: bytes | None = None,
) -> Path:
    try:
        parent_metadata = allowed_parent.lstat()
    except FileNotFoundError as exc:
        raise ArtifactValidationError(
            "allowed file directory is missing"
        ) from exc
    if allowed_parent.is_symlink() or not stat.S_ISDIR(
        parent_metadata.st_mode
    ):
        raise ArtifactValidationError(
            "allowed file directory must be a non-symlink directory"
        )
    allowed = allowed_parent.resolve(strict=True)
    if path.parent.resolve(strict=True) != allowed:
        raise ArtifactValidationError("file is outside its allowed directory")
    try:
        metadata = path.lstat()
    except FileNotFoundError as exc:
        raise ArtifactValidationError(f"{path.name} is missing") from exc
    if not stat.S_ISREG(metadata.st_mode) or path.is_symlink():
        raise ArtifactValidationError(
            f"{path.name} must be a non-symlink regular file"
        )
    resolved = path.resolve(strict=True)
    if resolved.parent != allowed:
        raise ArtifactValidationError("file resolves outside its allowed directory")
    if metadata.st_size < 1 or metadata.st_size > max_bytes:
        raise ArtifactValidationError(
            f"{path.name} has an invalid or excessive size"
        )
    if magic is not None:
        with path.open("rb") as handle:
            if not handle.read(len(magic)).startswith(magic):
                raise ArtifactValidationError(
                    f"{path.name} has invalid file content"
                )
    return resolved


def validate_json_artifact(path: Path, artifacts_dir: Path) -> Any:
    regular = require_regular_file(
        path,
        allowed_parent=artifacts_dir,
        max_bytes=MAX_JSON_BYTES,
    )
    try:
        return json.loads(
            regular.read_text(encoding="utf-8"),
            object_pairs_hook=_reject_duplicate_keys,
            parse_constant=lambda value: (_ for _ in ()).throw(
                ArtifactValidationError(
                    f"JSON contains invalid numeric value {value}"
                )
            ),
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ArtifactValidationError(
            f"{path.name} is not valid UTF-8 JSON"
        ) from exc


def _summary_value(row: dict[str, Any]) -> float:
    for key in SUMMARY_TOTAL_KEYS:
        value = row.get(key)
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)) and math.isfinite(float(value)):
            if float(value) < 0:
                break
            return float(value)
    raise ArtifactValidationError(
        "summary rows require a finite non-negative quantity/count total"
    )


def _validate_summary(
    rows: list[dict[str, Any]],
    *,
    dimension_keys: tuple[str, ...],
    expected: dict[tuple[str, ...], float],
    label: str,
) -> None:
    if expected and not rows:
        raise ArtifactValidationError(f"{label} summary is required")
    actual: dict[tuple[str, ...], float] = {}
    for row in rows:
        if len(row) > 50:
            raise ArtifactValidationError(
                f"{label} summary contains too many fields"
            )
        for key, value in row.items():
            if not isinstance(key, str) or len(key) > 128:
                raise ArtifactValidationError(
                    f"{label} summary contains an invalid field name"
                )
            if isinstance(value, str) and len(value) > 2_000:
                raise ArtifactValidationError(
                    f"{label} summary contains overlong text"
                )
            if isinstance(value, (list, dict)):
                raise ArtifactValidationError(
                    f"{label} summary values must be scalar"
                )
            if (
                isinstance(value, (int, float))
                and not isinstance(value, bool)
                and not math.isfinite(float(value))
            ):
                raise ArtifactValidationError(
                    f"{label} summary contains a non-finite number"
                )
        dimension_values: list[str] = []
        for key in dimension_keys:
            value = row.get(key)
            if not isinstance(value, str) or not value.strip():
                raise ArtifactValidationError(
                    f"{label} summary is missing dimension {key}"
                )
            dimension_values.append(value.strip())
        dimension = tuple(dimension_values)
        if dimension in actual:
            raise ArtifactValidationError(
                f"{label} summary has a duplicate dimension"
            )
        actual[dimension] = _summary_value(row)
    if set(actual) != set(expected):
        raise ArtifactValidationError(
            f"{label} summary dimensions do not match assets"
        )
    for dimension, total in expected.items():
        if not math.isclose(
            actual[dimension], total, rel_tol=1e-9, abs_tol=1e-6
        ):
            raise ArtifactValidationError(
                f"{label} total for {dimension} does not match assets"
            )


def _displayed_page_size(
    reader: PdfReader,
    page_number: int,
    *,
    label: str,
) -> tuple[float, float]:
    if page_number > len(reader.pages):
        raise ArtifactValidationError(
            f"{label} references a page outside the PDF"
        )
    width, height = displayed_size(reader.pages[page_number - 1])
    if (
        not math.isfinite(width)
        or not math.isfinite(height)
        or width <= 0
        or height <= 0
        or width > 10_000_000
        or height > 10_000_000
    ):
        raise ArtifactValidationError(
            "source PDF contains invalid page dimensions"
        )
    return width, height


def _validate_bbox_bounds(
    bbox: BoundingBox,
    *,
    width: float,
    height: float,
    label: str,
) -> None:
    if bbox.x1 > width or bbox.y1 > height:
        raise ArtifactValidationError(
            f"{label} bbox is outside its displayed PDF page"
        )


def _point_in_bbox(point: Point, bbox: BoundingBox) -> bool:
    return (
        bbox.x0 <= point.x <= bbox.x1
        and bbox.y0 <= point.y <= bbox.y1
    )


def _bboxes_intersect(left: BoundingBox, right: BoundingBox) -> bool:
    return not (
        left.x1 < right.x0
        or right.x1 < left.x0
        or left.y1 < right.y0
        or right.y1 < left.y0
    )


def _segment_intersects_bbox(
    start: Point,
    end: Point,
    bbox: BoundingBox,
) -> bool:
    if _point_in_bbox(start, bbox) or _point_in_bbox(end, bbox):
        return True

    delta_x = end.x - start.x
    delta_y = end.y - start.y
    lower = 0.0
    upper = 1.0
    for direction, offset in (
        (-delta_x, start.x - bbox.x0),
        (delta_x, bbox.x1 - start.x),
        (-delta_y, start.y - bbox.y0),
        (delta_y, bbox.y1 - start.y),
    ):
        if math.isclose(direction, 0.0, abs_tol=1e-12):
            if offset < 0:
                return False
            continue
        ratio = offset / direction
        if direction < 0:
            if ratio > upper:
                return False
            lower = max(lower, ratio)
        else:
            if ratio < lower:
                return False
            upper = min(upper, ratio)
    return lower <= upper


def _asset_intersects_bbox(
    asset: TakeoffAsset,
    bbox: BoundingBox,
) -> bool:
    if asset.bbox is not None:
        return _bboxes_intersect(asset.bbox, bbox)
    if asset.x is not None and asset.y is not None:
        return _point_in_bbox(Point(x=asset.x, y=asset.y), bbox)
    if asset.path:
        return any(
            _segment_intersects_bbox(left, right, bbox)
            for left, right in zip(asset.path, asset.path[1:])
        )
    return False


def validate_takeoff_artifact(
    takeoff_path: Path,
    *,
    drawings_path: Path,
    artifacts_dir: Path,
    inputs_dir: Path,
) -> tuple[TakeoffDocument, int]:
    require_regular_file(
        drawings_path,
        allowed_parent=inputs_dir,
        max_bytes=MAX_PDF_BYTES,
        magic=b"%PDF-",
    )
    raw = validate_json_artifact(takeoff_path, artifacts_dir)
    try:
        takeoff = TakeoffDocument.model_validate(raw)
    except Exception as exc:
        raise ArtifactValidationError(
            "takeoff.json does not satisfy the output schema"
        ) from exc

    reader = PdfReader(str(drawings_path), strict=True)
    actual_pages = len(reader.pages)
    if takeoff.source.page_count != actual_pages:
        raise ArtifactValidationError(
            "takeoff.json page_count does not match the uploaded PDF"
        )

    legend_boxes_by_page: dict[int, list[BoundingBox]] = defaultdict(list)
    for entry in takeoff.legend_entries:
        width, height = _displayed_page_size(
            reader,
            entry.page,
            label=f"legend entry {entry.legend_entry_id}",
        )
        _validate_bbox_bounds(
            entry.bbox,
            width=width,
            height=height,
            label=f"legend entry {entry.legend_entry_id}",
        )
        legend_boxes_by_page[entry.page].append(entry.bbox)

    for symbol in takeoff.unresolved_symbols:
        width, height = _displayed_page_size(
            reader,
            symbol.page,
            label=f"unresolved symbol {symbol.unresolved_symbol_id}",
        )
        _validate_bbox_bounds(
            symbol.bbox,
            width=width,
            height=height,
            label=f"unresolved symbol {symbol.unresolved_symbol_id}",
        )

    unit_ids: set[str] = set()
    by_code: dict[tuple[str, str, str, str, str], float] = defaultdict(
        float
    )
    by_area_code: dict[
        tuple[str, str, str, str, str, str],
        float,
    ] = defaultdict(float)
    for asset in takeoff.assets:
        if asset.unit_id in unit_ids:
            raise ArtifactValidationError(
                "takeoff.json contains duplicate unit_id values"
            )
        unit_ids.add(asset.unit_id)
        width, height = _displayed_page_size(
            reader,
            asset.page,
            label=asset.unit_id,
        )
        if asset.bbox is not None:
            _validate_bbox_bounds(
                asset.bbox,
                width=width,
                height=height,
                label=asset.unit_id,
            )
        elif asset.x is not None and asset.y is not None:
            if asset.x > width or asset.y > height:
                raise ArtifactValidationError(
                    f"{asset.unit_id} point is outside its displayed PDF page"
                )
        elif asset.path:
            for index, point in enumerate(asset.path):
                if point.x > width or point.y > height:
                    raise ArtifactValidationError(
                        f"{asset.unit_id} path point {index} is outside its "
                        "displayed PDF page"
                    )
            evidence = asset.scale_evidence
            if evidence is None:
                raise ArtifactValidationError(
                    f"{asset.unit_id} is missing scale evidence"
                )
            if evidence.page != asset.page or evidence.sheet != asset.sheet:
                raise ArtifactValidationError(
                    f"{asset.unit_id} scale evidence must come from the same "
                    "source page and sheet"
                )
            evidence_width, evidence_height = _displayed_page_size(
                reader,
                evidence.page,
                label=f"{asset.unit_id} scale evidence",
            )
            _validate_bbox_bounds(
                evidence.bbox,
                width=evidence_width,
                height=evidence_height,
                label=f"{asset.unit_id} scale evidence",
            )
            expected_quantity = (
                asset.display_path_length_points()
                * evidence.derived_real_units_per_pdf_point()
            )
            if not math.isclose(
                asset.quantity,
                expected_quantity,
                rel_tol=LINEAR_QUANTITY_REL_TOLERANCE,
                abs_tol=LINEAR_QUANTITY_ABS_TOLERANCE,
            ):
                raise ArtifactValidationError(
                    f"{asset.unit_id} quantity does not match deterministic "
                    "path-length times scale"
                )
        else:
            raise ArtifactValidationError(
                f"{asset.unit_id} has no supported geometry"
            )

        if any(
            _asset_intersects_bbox(asset, legend_box)
            for legend_box in legend_boxes_by_page.get(asset.page, [])
        ):
            raise ArtifactValidationError(
                f"{asset.unit_id} overlaps a legend exemplar and cannot enter "
                "takeoff totals"
            )

        by_code[
            (
                asset.legend_entry_id,
                asset.code,
                asset.description,
                asset.measurement_kind,
                asset.unit,
            )
        ] += asset.quantity
        by_area_code[
            (
                asset.area_code,
                asset.legend_entry_id,
                asset.code,
                asset.description,
                asset.measurement_kind,
                asset.unit,
            )
        ] += asset.quantity

    _validate_summary(
        takeoff.by_code,
        dimension_keys=(
            "legend_entry_id",
            "code",
            "description",
            "measurement_kind",
            "unit",
        ),
        expected=dict(by_code),
        label="by_code",
    )
    if takeoff.by_area and all(
        isinstance(row.get("area_code"), str) for row in takeoff.by_area
    ):
        _validate_summary(
            takeoff.by_area,
            dimension_keys=(
                "area_code",
                "legend_entry_id",
                "code",
                "description",
                "measurement_kind",
                "unit",
            ),
            expected=dict(by_area_code),
            label="by_area",
        )
    else:
        by_area: dict[
            tuple[str, str, str, str, str, str],
            float,
        ] = defaultdict(float)
        for asset in takeoff.assets:
            by_area[
                (
                    asset.area,
                    asset.legend_entry_id,
                    asset.code,
                    asset.description,
                    asset.measurement_kind,
                    asset.unit,
                )
            ] += asset.quantity
        _validate_summary(
            takeoff.by_area,
            dimension_keys=(
                "area",
                "legend_entry_id",
                "code",
                "description",
                "measurement_kind",
                "unit",
            ),
            expected=dict(by_area),
            label="by_area",
        )
    return takeoff, actual_pages


def reject_secret_material(path: Path, *, secret: str) -> None:
    if not secret:
        return
    encoded = secret.encode("utf-8")
    if path.suffix.lower() == ".xlsx":
        with zipfile.ZipFile(path) as archive:
            if any(encoded in archive.read(entry) for entry in archive.infolist()):
                raise ArtifactValidationError(
                    f"{path.name} contains credential material"
                )
        return
    if encoded in path.read_bytes():
        raise ArtifactValidationError(
            f"{path.name} contains credential material"
        )


def validate_xlsx_container(path: Path, allowed_parent: Path) -> Path:
    regular = require_regular_file(
        path,
        allowed_parent=allowed_parent,
        max_bytes=MAX_WORKBOOK_BYTES,
        magic=b"PK\x03\x04",
    )
    if not zipfile.is_zipfile(regular):
        raise ArtifactValidationError("takeoff.xlsx is not a valid ZIP/XLSX")
    tail_size = min(regular.stat().st_size, 65_557)
    with regular.open("rb") as handle:
        handle.seek(-tail_size, 2)
        tail = handle.read()
    marker = tail.rfind(b"PK\x05\x06")
    if marker < 0 or marker + 22 > len(tail):
        raise ArtifactValidationError("takeoff.xlsx has no valid ZIP trailer")
    comment_length = int.from_bytes(tail[marker + 20 : marker + 22], "little")
    if marker + 22 + comment_length != len(tail):
        raise ArtifactValidationError(
            "takeoff.xlsx contains trailing/polyglot content"
        )

    with zipfile.ZipFile(regular) as archive:
        entries = archive.infolist()
        if not entries or len(entries) > MAX_ZIP_ENTRIES:
            raise ArtifactValidationError(
                "takeoff.xlsx has an invalid ZIP entry count"
            )
        names = [entry.filename for entry in entries]
        if len(names) != len(set(names)):
            raise ArtifactValidationError(
                "takeoff.xlsx contains duplicate ZIP paths"
            )
        total_uncompressed = 0
        for entry in entries:
            parts = Path(entry.filename).parts
            if entry.flag_bits & 0x1 or entry.filename.startswith("/"):
                raise ArtifactValidationError(
                    "takeoff.xlsx contains encrypted or absolute ZIP entries"
                )
            if ".." in parts:
                raise ArtifactValidationError(
                    "takeoff.xlsx contains a traversal ZIP entry"
                )
            total_uncompressed += entry.file_size
            if total_uncompressed > MAX_ZIP_UNCOMPRESSED_BYTES:
                raise ArtifactValidationError(
                    "takeoff.xlsx expands beyond the allowed size"
                )
            if entry.compress_size == 0 and entry.file_size > 0:
                raise ArtifactValidationError(
                    "takeoff.xlsx has an invalid compression ratio"
                )
            if (
                entry.compress_size > 0
                and entry.file_size / entry.compress_size > 1_000
            ):
                raise ArtifactValidationError(
                    "takeoff.xlsx has an excessive compression ratio"
                )
            lowered = entry.filename.lower()
            if any(
                token in lowered
                for token in (
                    "vbaproject.bin",
                    "externallinks/",
                    "embeddings/",
                    "oleobjects/",
                    "connections.xml",
                    "querytables/",
                    "richdata/",
                    "datamodel/",
                )
            ):
                raise ArtifactValidationError(
                    "takeoff.xlsx contains active or external content"
                )
            if lowered.endswith((".xml", ".rels")):
                payload = archive.read(entry)
                lowered_payload = payload.lower()
                if lowered.endswith(".rels"):
                    _reject_external_relationships(payload)
                if lowered == "xl/workbook.xml":
                    _reject_defined_names(payload)
                if (
                    b"macrosheet" in lowered_payload
                    or b"macroenabled" in lowered_payload
                    or b"<dde" in lowered_payload
                    or b"activex" in lowered_payload
                    or b"_xlnm.auto_open" in lowered_payload
                    or b"_xlnm.auto_close" in lowered_payload
                ):
                    raise ArtifactValidationError(
                        "takeoff.xlsx contains external or active relationships"
                    )
    return regular


def _normalized_header(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _audit_rows(workbook: Any) -> tuple[dict[str, int], list[list[Any]]]:
    sheet = next(
        (
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.title.strip().lower() == "takeoff"
        ),
        None,
    )
    if sheet is None:
        raise ArtifactValidationError(
            "takeoff.xlsx requires a sheet named Takeoff"
        )
    header_map: dict[str, int] | None = None
    data: list[list[Any]] = []
    for row_number, cells in enumerate(
        sheet.iter_rows(values_only=True), start=1
    ):
        values = list(cells)
        if header_map is None:
            if row_number > 20:
                break
            normalized = [_normalized_header(value) for value in values]
            if "unit_id" not in normalized:
                continue
            if len([value for value in normalized if value]) != len(
                set(value for value in normalized if value)
            ):
                raise ArtifactValidationError(
                    "Takeoff sheet contains duplicate headers"
                )
            header_map = {
                value: index
                for index, value in enumerate(normalized)
                if value
            }
            missing = set(REQUIRED_TAKEOFF_HEADERS) - set(header_map)
            if missing:
                raise ArtifactValidationError(
                    "Takeoff sheet is missing required headers: "
                    + ", ".join(sorted(missing))
                )
            continue
        if any(value is not None and value != "" for value in values):
            data.append(values)
    if header_map is None:
        raise ArtifactValidationError(
            "Takeoff sheet has no machine-readable header row"
        )
    return header_map, data


def validate_workbook_artifact(
    workbook_path: Path,
    *,
    takeoff: TakeoffDocument,
    artifacts_dir: Path,
) -> None:
    regular = validate_xlsx_container(workbook_path, artifacts_dir)
    try:
        workbook = load_workbook(
            regular,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ArtifactValidationError(
            "takeoff.xlsx cannot be parsed safely"
        ) from exc
    try:
        if not workbook.worksheets or len(workbook.worksheets) > MAX_WORKBOOK_SHEETS:
            raise ArtifactValidationError(
                "takeoff.xlsx has an invalid sheet count"
            )
        total_cells = 0
        for worksheet in workbook.worksheets:
            max_row = worksheet.max_row or 0
            max_column = worksheet.max_column or 0
            if (
                max_row > MAX_WORKBOOK_ROWS
                or max_column > 16_384
            ):
                raise ArtifactValidationError(
                    "takeoff.xlsx exceeds worksheet bounds"
                )
            total_cells += max_row * max_column
            if total_cells > MAX_WORKBOOK_CELLS:
                raise ArtifactValidationError(
                    "takeoff.xlsx exceeds the cell inspection limit"
                )
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and len(value) > MAX_CELL_TEXT:
                        raise ArtifactValidationError(
                            "takeoff.xlsx contains overlong cell text"
                        )
                    if cell.data_type == "e":
                        raise ArtifactValidationError(
                            "takeoff.xlsx contains an obvious formula error"
                        )
                    if (
                        cell.data_type == "f"
                        or isinstance(value, str)
                        and value.startswith("=")
                    ):
                        raise ArtifactValidationError(
                            "takeoff.xlsx contains a formula; formulas are not "
                            "permitted"
                        )

        headers, rows = _audit_rows(workbook)
        expected = {asset.unit_id: asset for asset in takeoff.assets}
        observed: dict[str, list[Any]] = {}
        for row in rows:
            unit_id = row[headers["unit_id"]]
            if not isinstance(unit_id, str) or not unit_id.strip():
                raise ArtifactValidationError(
                    "Takeoff sheet unit_id values must be text"
                )
            unit_id = unit_id.strip()
            if unit_id in observed:
                raise ArtifactValidationError(
                    "Takeoff sheet contains duplicate unit_id values"
                )
            observed[unit_id] = row
        if set(observed) != set(expected):
            raise ArtifactValidationError(
                "Takeoff sheet unit IDs do not match takeoff.json"
            )

        text_fields = (
            "legend_entry_id",
            "measurement_kind",
            "code",
            "description",
            "sheet",
            "area_code",
            "area",
            "level",
            "method",
            "confidence",
            "unit",
        )
        for unit_id, asset in expected.items():
            row = observed[unit_id]
            page = row[headers["page"]]
            quantity = row[headers["quantity"]]
            if isinstance(page, bool) or not isinstance(page, (int, float)):
                raise ArtifactValidationError(
                    f"Takeoff sheet page is invalid for {unit_id}"
                )
            if int(page) != page or int(page) != asset.page:
                raise ArtifactValidationError(
                    f"Takeoff sheet page differs for {unit_id}"
                )
            if (
                isinstance(quantity, bool)
                or not isinstance(quantity, (int, float))
                or not math.isfinite(float(quantity))
                or not math.isclose(
                    float(quantity),
                    asset.quantity,
                    rel_tol=1e-9,
                    abs_tol=1e-6,
                )
            ):
                raise ArtifactValidationError(
                    f"Takeoff sheet quantity differs for {unit_id}"
                )
            for field in text_fields:
                value = row[headers[field]]
                if not isinstance(value, str) or value.strip() != str(
                    getattr(asset, field)
                ):
                    raise ArtifactValidationError(
                        f"Takeoff sheet {field} differs for {unit_id}"
                    )
            linear_fields = (
                "path_length_pdf_points",
                "scale_kind",
                "scale_source_page",
                "scale_source_sheet",
                "scale_source_text",
                "scale_real_units_per_pdf_point",
            )
            if asset.measurement_kind == "count":
                if any(
                    row[headers[field]] not in (None, "")
                    for field in linear_fields
                ):
                    raise ArtifactValidationError(
                        f"Takeoff sheet count row contains linear evidence "
                        f"for {unit_id}"
                    )
                continue

            evidence = asset.scale_evidence
            if evidence is None:
                raise ArtifactValidationError(
                    f"Takeoff sheet linear evidence is unavailable for "
                    f"{unit_id}"
                )
            numeric_fields = {
                "path_length_pdf_points": (
                    asset.display_path_length_points()
                ),
                "scale_real_units_per_pdf_point": (
                    evidence.derived_real_units_per_pdf_point()
                ),
            }
            for field, expected_value in numeric_fields.items():
                value = row[headers[field]]
                if (
                    isinstance(value, bool)
                    or not isinstance(value, (int, float))
                    or not math.isfinite(float(value))
                    or not math.isclose(
                        float(value),
                        expected_value,
                        rel_tol=LINEAR_QUANTITY_REL_TOLERANCE,
                        abs_tol=LINEAR_QUANTITY_ABS_TOLERANCE,
                    )
                ):
                    raise ArtifactValidationError(
                        f"Takeoff sheet {field} differs for {unit_id}"
                    )
            scale_page = row[headers["scale_source_page"]]
            if (
                isinstance(scale_page, bool)
                or not isinstance(scale_page, (int, float))
                or int(scale_page) != scale_page
                or int(scale_page) != evidence.page
            ):
                raise ArtifactValidationError(
                    f"Takeoff sheet scale_source_page differs for {unit_id}"
                )
            expected_scale_text = {
                "scale_kind": evidence.kind,
                "scale_source_sheet": evidence.sheet,
                "scale_source_text": evidence.source_text,
            }
            for field, expected_value in expected_scale_text.items():
                value = row[headers[field]]
                if (
                    not isinstance(value, str)
                    or value.strip() != str(expected_value)
                ):
                    raise ArtifactValidationError(
                        f"Takeoff sheet {field} differs for {unit_id}"
                    )
    finally:
        workbook.close()

    try:
        cached_workbook = load_workbook(
            regular,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ArtifactValidationError(
            "takeoff.xlsx cached values cannot be parsed safely"
        ) from exc
    try:
        for worksheet in cached_workbook.worksheets:
            for row in worksheet.iter_rows():
                if any(cell.data_type == "e" for cell in row):
                    raise ArtifactValidationError(
                        "takeoff.xlsx contains a cached formula error"
                    )
    finally:
        cached_workbook.close()


def validate_pdf_artifact(
    path: Path,
    *,
    artifacts_dir: Path,
    expected_pages: int,
) -> None:
    regular = require_regular_file(
        path,
        allowed_parent=artifacts_dir,
        max_bytes=MAX_PDF_BYTES,
        magic=b"%PDF-",
    )
    try:
        pages = len(PdfReader(str(regular), strict=True).pages)
    except Exception as exc:
        raise ArtifactValidationError(
            f"{path.name} is not a parseable PDF"
        ) from exc
    if pages != expected_pages:
        raise ArtifactValidationError(
            f"{path.name} page count does not match the source"
        )
