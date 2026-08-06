from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from pypdf import PdfWriter

from app.config import Settings
from app.codex_runner import CodexRunOutcome
from app.models import JobRecord, JobStatus, RequestedScope
from app.pipeline import PipelineManager
from app.store import JobStore


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_replay_pipeline_creates_audited_pdf(tmp_path: Path) -> None:
    settings = Settings(
        data_dir=tmp_path / "data",
        codex_bin=sys.executable,
        default_model="gpt-5.6-sol",
        max_upload_bytes=10_000_000,
        max_total_upload_bytes=10_000_000,
        service_api_token=None,
        max_workers=1,
        environment="test",
    )
    store = JobStore(settings.data_dir / "jobs")
    record = JobRecord(
        id="replayjob",
        status=JobStatus.queued,
        model=settings.default_model,
        requested_scopes=[
            RequestedScope.fixture_counts,
            RequestedScope.cable_runs,
        ],
    )
    job_dir = store.create(record)
    source_pdf = job_dir / "inputs" / "drawings.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=100)
    writer.write(str(source_pdf))

    replay = job_dir / "inputs" / "replay_takeoff.json"
    replay.write_text(
        json.dumps(
            {
                "source": {
                    "sha256": sha256(source_pdf),
                    "page_count": 1
                },
                "legend_entries": [
                    {
                        "legend_entry_id": "LEGEND-TC",
                        "code": "TC",
                        "description": "Test receptacle",
                        "page": 1,
                        "sheet": "E-001",
                        "bbox": {
                            "x0": 5,
                            "y0": 5,
                            "x1": 15,
                            "y1": 15
                        }
                    },
                    {
                        "legend_entry_id": "LEGEND-CBL",
                        "code": "CBL",
                        "description": "Type C cable",
                        "page": 1,
                        "sheet": "E-001",
                        "bbox": {
                            "x0": 20,
                            "y0": 5,
                            "x1": 30,
                            "y1": 15
                        }
                    }
                ],
                "assets": [
                    {
                        "unit_id": "TEST-TC-001-0001",
                        "legend_entry_id": "LEGEND-TC",
                        "measurement_kind": "count",
                        "code": "TC",
                        "description": "Test receptacle",
                        "page": 1,
                        "sheet": "E-101",
                        "area_code": "L1-A",
                        "area": "Level 1 - Area A",
                        "level": "1",
                        "method": "counted",
                        "confidence": "medium",
                        "x": 75,
                        "y": 40
                    },
                    {
                        "unit_id": "TEST-CBL-001-0001",
                        "legend_entry_id": "LEGEND-CBL",
                        "measurement_kind": "linear",
                        "code": "CBL",
                        "description": "Type C cable",
                        "page": 1,
                        "sheet": "E-101",
                        "area_code": "L1-A",
                        "area": "Level 1 - Area A",
                        "level": "1",
                        "method": "scaled centerline",
                        "confidence": "high",
                        "path": [
                            {"x": 20, "y": 80},
                            {"x": 80, "y": 80}
                        ],
                        "scale_evidence": {
                            "kind": "calibrated_dimension",
                            "page": 1,
                            "sheet": "E-101",
                            "bbox": {
                                "x0": 160,
                                "y0": 5,
                                "x1": 195,
                                "y1": 20
                            },
                            "source_text": "15 ft calibration",
                            "real_units_per_pdf_point": 0.5,
                            "unit": "ft",
                            "calibration": {
                                "start": {"x": 160, "y": 10},
                                "end": {"x": 190, "y": 10},
                                "known_length": 15,
                                "unit": "ft"
                            }
                        },
                        "quantity": 30,
                        "unit": "ft"
                    }
                ],
                "unresolved_symbols": [
                    {
                        "unresolved_symbol_id": "UNRESOLVED-001",
                        "page": 1,
                        "sheet": "E-101",
                        "bbox": {
                            "x0": 120,
                            "y0": 60,
                            "x1": 130,
                            "y1": 70
                        },
                        "visible_label": "?",
                        "reason": "No defensible legend mapping.",
                        "confidence": "low"
                    }
                ],
                "by_code": [
                    {
                        "legend_entry_id": "LEGEND-TC",
                        "code": "TC",
                        "description": "Test receptacle",
                        "measurement_kind": "count",
                        "unit": "EA",
                        "quantity": 1
                    },
                    {
                        "legend_entry_id": "LEGEND-CBL",
                        "code": "CBL",
                        "description": "Type C cable",
                        "measurement_kind": "linear",
                        "unit": "ft",
                        "quantity": 30
                    }
                ],
                "by_area": [
                    {
                        "area_code": "L1-A",
                        "legend_entry_id": "LEGEND-TC",
                        "code": "TC",
                        "description": "Test receptacle",
                        "measurement_kind": "count",
                        "unit": "EA",
                        "quantity": 1
                    },
                    {
                        "area_code": "L1-A",
                        "legend_entry_id": "LEGEND-CBL",
                        "code": "CBL",
                        "description": "Type C cable",
                        "measurement_kind": "linear",
                        "unit": "ft",
                        "quantity": 30
                    }
                ],
                "limitations": []
            }
        ),
        encoding="utf-8",
    )

    manager = PipelineManager(settings, store)
    manager._run("replayjob", None, replay, None)
    completed = store.load("replayjob")

    assert completed.status == JobStatus.completed
    assert completed.metrics["annotated_units"] == 2
    assert completed.metrics["legend_entries"] == 2
    assert completed.metrics["mapped_assets"] == 2
    assert completed.metrics["unresolved_symbols"] == 1
    assert completed.metrics["legend_coverage_percent"] == pytest.approx(
        66.6666666667
    )
    assert completed.metrics["count_placements"] == 1
    assert completed.metrics["linear_runs"] == 1
    assert completed.metrics["linear_path_points"] == 2
    assert completed.metrics["linear_quantity_by_unit"] == {"ft": 30}
    assert completed.processor_usage is None
    assert "estimated_cost_usd" not in completed.metrics
    assert "annotated_drawings.pdf" in completed.artifacts
    assert "takeoff.xlsx" in completed.artifacts
    assert (
        job_dir / "artifacts" / "annotated_drawings.pdf"
    ).stat().st_size > source_pdf.stat().st_size
    workbook = load_workbook(
        job_dir / "artifacts" / "takeoff.xlsx",
        read_only=True,
        keep_links=False,
    )
    try:
        assert workbook.sheetnames[:2] == [
            "Resumen Takeoff",
            "Codigos Elementos",
        ]
        assert workbook.sheetnames[-1] == "Takeoff"
        assert workbook["Takeoff"].sheet_state == "hidden"
    finally:
        workbook.close()


def test_pipeline_keeps_processor_usage_out_of_public_metrics(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    settings = Settings(
        data_dir=tmp_path / "data",
        codex_bin=sys.executable,
        default_model="gpt-5.6-sol",
        max_upload_bytes=10_000_000,
        max_total_upload_bytes=10_000_000,
        service_api_token=None,
        max_workers=1,
        environment="test",
    )
    store = JobStore(settings.data_dir / "jobs")
    job_dir = store.create(
        JobRecord(
            id="usagejob",
            status=JobStatus.queued,
            model=settings.default_model,
        )
    )
    drawings = job_dir / "inputs" / "drawings.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=100)
    writer.write(str(drawings))
    processor_usage = {
        "schema_version": 1,
        "provider": "openai",
        "model": "gpt-5.6-sol",
        "pricing_as_of": "2026-08-06",
        "currency": "USD",
        "usage_turns": 1,
        "input_tokens": 100_000,
        "uncached_input_tokens": 70_000,
        "cached_input_tokens": 20_000,
        "cache_write_tokens": 10_000,
        "output_tokens": 5_000,
        "reasoning_output_tokens": 1_000,
        "estimated_cost_usd": 0.5725,
        "estimated_cost_usd_upper_bound": None,
        "estimated_cost_usd_all_input_uncached": 0.65,
        "estimated_cost_usd_all_input_uncached_upper_bound": None,
        "long_context_pricing_may_apply": False,
        "rate_snapshot_usd_per_million": {
            "input": 5,
            "cached_input": 0.5,
            "cache_write": 6.25,
            "output": 30,
        },
    }
    drawing_validation = SimpleNamespace(
        validator_exit_code=0,
        source_count=1,
        page_count=1,
        pending_pages=0,
        text_reviewed_pages=0,
        visually_reviewed_pages=1,
        image_only_pages=0,
        unknown_revision_pages=1,
        object_count=1,
        fact_count=1,
        evidence_count=1,
        low_confidence_facts=0,
        unverified_facts=0,
        open_references=0,
        open_conflicts=0,
        warnings=("Open reference remains unresolved",),
    )

    def fake_prepare_drawing_index(**kwargs: object) -> SimpleNamespace:
        current_job = Path(str(kwargs["job_dir"]))
        return SimpleNamespace(
            index_dir=current_job / "work" / "drawing-index",
            skill_dir=(
                current_job
                / ".agents"
                / "skills"
                / "analyze-building-drawings"
            ),
            skill_sha256="a" * 64,
            source_sha256=str(kwargs["expected_sha256"]),
            validation=drawing_validation,
        )

    def fake_run_codex(**_kwargs: object) -> CodexRunOutcome:
        assert _kwargs["analysis_skill_sha256"] == "a" * 64
        assert str(_kwargs["analysis_profile"].value) == (
            "analyze-building-drawings@2026-08-06"
        )
        artifacts = Path(str(_kwargs["job_dir"])) / "artifacts"
        (artifacts / "takeoff.json").write_text("{}", encoding="utf-8")
        (artifacts / "methodology.json").write_text("{}", encoding="utf-8")
        (artifacts / "takeoff.xlsx").write_bytes(b"PK-test")
        return CodexRunOutcome(result={"status": "completed"}, metrics=processor_usage)

    def fake_annotate(
        _source: Path, output: Path, _assets: object, **_kwargs: object
    ) -> SimpleNamespace:
        output.write_bytes(b"%PDF-1.4\n%%EOF\n")
        return SimpleNamespace(
            annotated_asset_count=0,
            skipped_asset_count=0,
            model_dump_json=lambda **_kwargs: "{}",
        )

    fake_takeoff = SimpleNamespace(
        source=SimpleNamespace(sha256=sha256(drawings)),
        assets=[],
        unresolved_symbols=[],
        legend_entries=[],
    )
    monkeypatch.setattr("app.pipeline.run_codex", fake_run_codex)
    monkeypatch.setattr(
        "app.pipeline.prepare_drawing_index",
        fake_prepare_drawing_index,
    )
    monkeypatch.setattr(
        "app.pipeline.validate_drawing_index",
        lambda *_args, **_kwargs: drawing_validation,
    )
    monkeypatch.setattr(
        "app.pipeline.validate_takeoff_index_alignment",
        lambda *_args, **_kwargs: SimpleNamespace(
            legend_objects=0,
            asset_objects=0,
            quantity_facts=0,
            instance_relationships=0,
        ),
    )
    monkeypatch.setattr(
        "app.pipeline.validate_takeoff_artifact",
        lambda *_args, **_kwargs: (fake_takeoff, 1),
    )
    monkeypatch.setattr(
        "app.pipeline.validate_workbook_artifact",
        lambda *_args, **_kwargs: None,
    )
    built_workbooks: list[Path] = []

    def fake_build_workbook(_takeoff: object, path: Path) -> None:
        built_workbooks.append(path)
        path.write_bytes(b"PK-test")

    monkeypatch.setattr("app.pipeline.build_takeoff_workbook", fake_build_workbook)
    monkeypatch.setattr(
        "app.pipeline.validate_json_artifact",
        lambda *_args, **_kwargs: {},
    )
    monkeypatch.setattr(
        "app.pipeline.reject_secret_material",
        lambda *_args, **_kwargs: None,
    )
    monkeypatch.setattr("app.pipeline.annotate_pdf", fake_annotate)
    monkeypatch.setattr(
        "app.pipeline.validate_pdf_artifact",
        lambda *_args, **_kwargs: None,
    )

    PipelineManager(settings, store)._run("usagejob", "secret", None, None)
    completed = store.load("usagejob")

    assert completed.status == JobStatus.completed
    assert completed.processor_usage is not None
    assert completed.processor_usage.estimated_cost_usd == 0.5725
    assert (
        completed.processor_usage.estimated_cost_usd_all_input_uncached
        == 0.65
    )
    assert "estimated_cost_usd" not in completed.metrics
    assert "processor_usage" not in completed.metrics
    assert completed.metrics["analysis_profile"] == (
        "analyze-building-drawings@2026-08-06"
    )
    assert completed.metrics["drawing_index"]["pages"] == 1
    assert completed.metrics["drawing_index_alignment"]["asset_objects"] == 0
    assert built_workbooks == [job_dir / "artifacts" / "takeoff.xlsx"]
    methodology = json.loads(
        (job_dir / "artifacts" / "methodology.json").read_text(
            encoding="utf-8"
        )
    )
    assert methodology["analysis_skill"]["index_validation"]["warnings"] == [
        "Open reference remains unresolved"
    ]

    failed_dir = store.create(
        JobRecord(
            id="usagefailure",
            status=JobStatus.queued,
            model=settings.default_model,
        )
    )
    failed_writer = PdfWriter()
    failed_writer.add_blank_page(width=200, height=100)
    failed_writer.write(str(failed_dir / "inputs" / "drawings.pdf"))
    monkeypatch.setattr(
        "app.pipeline.validate_takeoff_artifact",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            ValueError("post-model validation failed")
        ),
    )

    PipelineManager(settings, store)._run(
        "usagefailure", "secret", None, None
    )
    failed = store.load("usagefailure")

    assert failed.status == JobStatus.failed
    assert failed.processor_usage is not None
    assert failed.processor_usage.estimated_cost_usd == 0.5725
    assert failed.processor_usage.estimated_cost_usd_all_input_uncached == 0.65
    assert "estimated_cost_usd" not in failed.metrics


def test_pipeline_keeps_diagnostics_private_and_public_error_generic(
    tmp_path: Path,
) -> None:
    settings = Settings(
        data_dir=tmp_path / "data",
        codex_bin=sys.executable,
        default_model="gpt-5.6-sol",
        max_upload_bytes=10_000_000,
        max_total_upload_bytes=10_000_000,
        service_api_token=None,
        max_workers=1,
        environment="test",
    )
    store = JobStore(settings.data_dir / "jobs")
    job_dir = store.create(
        JobRecord(
            id="failingjob",
            status=JobStatus.queued,
            model=settings.default_model,
        )
    )
    manager = PipelineManager(settings, store)

    manager._run("failingjob", None, None, None)

    failed = store.load("failingjob")
    assert failed.status == JobStatus.failed
    assert failed.error_code == "processing_failed"
    assert failed.retriable is True
    assert "drawings.pdf" not in failed.error
    diagnostic = job_dir / "work" / "pipeline-error.log"
    assert diagnostic.is_file()
    assert "drawings.pdf is missing" in diagnostic.read_text()
    assert "pipeline-error.log" not in failed.artifacts
