from __future__ import annotations

import copy
import hashlib
import json
import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from pypdf import PdfReader, PdfWriter
from pypdf.errors import PdfReadError
from pypdf.generic import NameObject

from app.config import Settings
from app.models import JobRecord, JobStatus, TakeoffDocument
from app.pipeline import PipelineManager
from app.store import JobStore
from app.validation import (
    ArtifactValidationError,
    validate_json_artifact,
    reject_secret_material,
    validate_takeoff_artifact,
    validate_workbook_artifact,
)


def make_source_pdf(path: Path) -> str:
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=100)
    writer.write(str(path))
    return hashlib.sha256(path.read_bytes()).hexdigest()


def make_duplicate_page_mode_pdf(path: Path) -> str:
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=100)
    writer.root_object.update(
        {
            NameObject("/PageMode"): NameObject("/UseNone"),
            NameObject("/PageMope"): NameObject("/UseNone"),
        }
    )
    writer.write(str(path))
    payload = path.read_bytes()
    placeholder = b"/PageMope /UseNone"
    assert payload.count(placeholder) == 1
    path.write_bytes(
        payload.replace(placeholder, b"/PageMode /UseNone", 1)
    )
    return hashlib.sha256(path.read_bytes()).hexdigest()


def takeoff_payload(source_hash: str) -> dict[str, object]:
    return {
        "source": {"sha256": source_hash, "page_count": 1},
        "legend_entries": [
            {
                "legend_entry_id": "LEGEND-DOOR",
                "code": "DOOR",
                "description": "Single door",
                "page": 1,
                "sheet": "A-001",
                "bbox": {"x0": 5, "y0": 5, "x1": 15, "y1": 15},
            }
        ],
        "assets": [
            {
                "unit_id": "TEST-DOOR-001",
                "legend_entry_id": "LEGEND-DOOR",
                "measurement_kind": "count",
                "code": "DOOR",
                "description": "Single door",
                "page": 1,
                "sheet": "A-101",
                "area_code": "L1",
                "area": "Level 1",
                "level": "1",
                "method": "symbol count",
                "confidence": "high",
                "x": 50,
                "y": 40,
                "quantity": 1,
                "unit": "EA",
            }
        ],
        "unresolved_symbols": [],
        "by_code": [
            {
                "legend_entry_id": "LEGEND-DOOR",
                "code": "DOOR",
                "description": "Single door",
                "measurement_kind": "count",
                "unit": "EA",
                "quantity": 1,
            }
        ],
        "by_area": [
            {
                "area_code": "L1",
                "legend_entry_id": "LEGEND-DOOR",
                "code": "DOOR",
                "description": "Single door",
                "measurement_kind": "count",
                "unit": "EA",
                "quantity": 1,
            }
        ],
        "limitations": [],
    }


def make_workbook(path: Path, payload: dict[str, object]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Takeoff"
    headers = [
        "unit_id",
        "legend_entry_id",
        "measurement_kind",
        "code",
        "description",
        "page",
        "sheet",
        "area_code",
        "area",
        "level",
        "method",
        "confidence",
        "quantity",
        "unit",
        "path_length_pdf_points",
        "scale_kind",
        "scale_source_page",
        "scale_source_sheet",
        "scale_source_text",
        "scale_real_units_per_pdf_point",
    ]
    worksheet.append(headers)
    asset = payload["assets"][0]
    scale = asset.get("scale_evidence") or {}
    row = {
        **asset,
        "path_length_pdf_points": (
            80 if asset.get("measurement_kind") == "linear" else None
        ),
        "scale_kind": scale.get("kind"),
        "scale_source_page": scale.get("page"),
        "scale_source_sheet": scale.get("sheet"),
        "scale_source_text": scale.get("source_text"),
        "scale_real_units_per_pdf_point": scale.get(
            "real_units_per_pdf_point"
        ),
    }
    worksheet.append([row.get(header) for header in headers])
    workbook.save(path)
    workbook.close()


def linear_takeoff_payload(source_hash: str) -> dict[str, object]:
    return {
        "source": {"sha256": source_hash, "page_count": 1},
        "legend_entries": [
            {
                "legend_entry_id": "LEGEND-CBL",
                "code": "CBL",
                "description": "Type C cable",
                "page": 1,
                "sheet": "E-101",
                "bbox": {"x0": 5, "y0": 5, "x1": 15, "y1": 15},
            }
        ],
        "assets": [
            {
                "unit_id": "TEST-CBL-001",
                "legend_entry_id": "LEGEND-CBL",
                "measurement_kind": "linear",
                "code": "CBL",
                "description": "Type C cable",
                "page": 1,
                "sheet": "E-101",
                "area_code": "L1",
                "area": "Level 1",
                "level": "1",
                "method": "scaled centerline",
                "confidence": "high",
                "path": [
                    {"x": 20, "y": 60},
                    {"x": 80, "y": 60},
                    {"x": 80, "y": 80},
                ],
                "scale_evidence": {
                    "kind": "calibrated_dimension",
                    "page": 1,
                    "sheet": "E-101",
                    "bbox": {
                        "x0": 160,
                        "y0": 5,
                        "x1": 195,
                        "y1": 20,
                    },
                    "source_text": "15 ft calibration",
                    "real_units_per_pdf_point": 0.5,
                    "unit": "ft",
                    "calibration": {
                        "start": {"x": 160, "y": 10},
                        "end": {"x": 190, "y": 10},
                        "known_length": 15,
                        "unit": "ft",
                    },
                },
                "quantity": 40,
                "unit": "ft",
            }
        ],
        "unresolved_symbols": [
            {
                "unresolved_symbol_id": "UNRESOLVED-001",
                "page": 1,
                "sheet": "E-101",
                "bbox": {"x0": 120, "y0": 60, "x1": 130, "y1": 70},
                "visible_label": "?",
                "candidate_code": "UNKNOWN",
                "reason": "No defensible legend mapping.",
                "confidence": "low",
            }
        ],
        "by_code": [
            {
                "legend_entry_id": "LEGEND-CBL",
                "code": "CBL",
                "description": "Type C cable",
                "measurement_kind": "linear",
                "unit": "ft",
                "quantity": 40,
            }
        ],
        "by_area": [
            {
                "area_code": "L1",
                "legend_entry_id": "LEGEND-CBL",
                "code": "CBL",
                "description": "Type C cable",
                "measurement_kind": "linear",
                "unit": "ft",
                "quantity": 40,
            }
        ],
        "limitations": ["One symbol remains unresolved and is excluded."],
    }


def add_external_relationship(path: Path) -> None:
    replacement = path.with_name("rewritten-takeoff.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement, "w"
    ) as destination:
        for entry in source.infolist():
            payload = source.read(entry)
            if entry.filename == "_rels/.rels":
                xml = payload.decode("utf-8")
                external = (
                    "<Relationship Id='rIdAuditExternal' "
                    "Type='http://schemas.openxmlformats.org/"
                    "officeDocument/2006/relationships/hyperlink' "
                    "Target='https://attacker.invalid/' "
                    "TargetMode='External'/>"
                )
                xml = xml.replace(
                    "</Relationships>",
                    f"{external}</Relationships>",
                )
                payload = xml.encode("utf-8")
            destination.writestr(entry, payload)
    replacement.replace(path)


def rewrite_workbook_definition_as_utf16(path: Path) -> None:
    replacement = path.with_name("utf16-takeoff.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement, "w"
    ) as destination:
        for entry in source.infolist():
            payload = source.read(entry)
            if entry.filename == "xl/workbook.xml":
                payload = payload.decode("utf-8").encode("utf-16")
            destination.writestr(entry, payload)
    replacement.replace(path)


def remove_worksheet_dimensions(path: Path) -> None:
    replacement = path.with_name("dimensionless-takeoff.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement, "w"
    ) as destination:
        for entry in source.infolist():
            payload = source.read(entry)
            if entry.filename.startswith(
                "xl/worksheets/"
            ) and entry.filename.endswith(".xml"):
                xml = payload.decode("utf-8")
                xml = re.sub(r"<dimension\b[^>]*/>", "", xml, flags=re.I)
                payload = xml.encode("utf-8")
            destination.writestr(entry, payload)
    replacement.replace(path)
    with zipfile.ZipFile(path, "r") as archive:
        worksheet_parts = [
            entry
            for entry in archive.namelist()
            if entry.startswith("xl/worksheets/") and entry.endswith(".xml")
        ]
        assert worksheet_parts
        assert all(
            re.search(
                rb"<(?:\w+:)?dimension\b",
                archive.read(entry),
                flags=re.I,
            )
            is None
            for entry in worksheet_parts
        )


def test_takeoff_and_workbook_reconcile(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)

    document, pages = validate_takeoff_artifact(
        takeoff_path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )
    validate_workbook_artifact(
        workbook_path, takeoff=document, artifacts_dir=artifacts
    )
    assert pages == 1


def test_codex_compatibility_normalizes_only_known_redundant_shapes(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    payload["legend_entries"][0]["method"] = "explicit"
    payload["legend_entries"][0]["confidence"] = "high"
    asset = payload["assets"][0]
    asset["geometry"] = {
        "bbox": {"x0": 45, "y0": 35, "x1": 55, "y1": 45}
    }
    asset.pop("x")
    asset.pop("y")
    payload["unresolved_symbols"] = [
        {
            "page": 1,
            "sheet": "A-101",
            "bbox": {"x0": 60, "y0": 60, "x1": 70, "y1": 70},
            "visible_label": "Unknown symbol",
            "reason": "No legend match",
            "confidence": "low",
        }
    ]
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(
        ArtifactValidationError,
        match="does not satisfy the output schema",
    ):
        validate_takeoff_artifact(
            takeoff_path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )

    document, _pages = validate_takeoff_artifact(
        takeoff_path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
        allow_codex_compatibility=True,
    )

    assert document.legend_entries[0].model_extra is None
    assert document.assets[0].bbox is not None
    assert document.assets[0].model_extra == {}
    assert document.unresolved_symbols[0].unresolved_symbol_id == (
        "UNRESOLVED-SERVER-0001"
    )
    assert json.loads(takeoff_path.read_text(encoding="utf-8")) == payload


def test_codex_compatibility_does_not_discard_unknown_shape_variants(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    payload["legend_entries"][0]["method"] = "inferred"
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(
        ArtifactValidationError,
        match="does not satisfy the output schema",
    ):
        validate_takeoff_artifact(
            takeoff_path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
            allow_codex_compatibility=True,
        )


def test_takeoff_validation_accepts_duplicate_page_mode(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_duplicate_page_mode_pdf(drawings))
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PdfReadError, match=r"key /PageMode"):
        len(PdfReader(str(drawings), strict=True).pages)

    document, pages = validate_takeoff_artifact(
        takeoff_path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )

    assert pages == 1
    assert len(document.assets) == 1


def test_unresolved_symbols_are_explicit_and_excluded_from_totals(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = linear_takeoff_payload(make_source_pdf(drawings))
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")

    document, _pages = validate_takeoff_artifact(
        takeoff_path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )

    assert len(document.assets) == 1
    assert document.by_code == [
        {
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "linear",
            "unit": "ft",
            "quantity": 40,
        }
    ]
    assert len(document.unresolved_symbols) == 1
    assert document.unresolved_symbols[0].candidate_code == "UNKNOWN"


def test_legend_exemplar_cannot_enter_takeoff_totals(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    payload["assets"][0]["x"] = 10
    payload["assets"][0]["y"] = 10
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(
        ArtifactValidationError,
        match="overlaps a legend exemplar",
    ):
        validate_takeoff_artifact(
            path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )


def test_linear_path_scale_and_workbook_evidence_reconcile(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = linear_takeoff_payload(make_source_pdf(drawings))
    takeoff_path = artifacts / "takeoff.json"
    takeoff_path.write_text(json.dumps(payload), encoding="utf-8")
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)

    document, _pages = validate_takeoff_artifact(
        takeoff_path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )
    validate_workbook_artifact(
        workbook_path,
        takeoff=document,
        artifacts_dir=artifacts,
    )

    run = document.assets[0]
    assert run.display_path_length_points() == 80
    assert run.quantity == 40
    assert run.unit == "ft"


def test_calibrated_scale_factor_is_independently_derived() -> None:
    payload = linear_takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    evidence = document.assets[0].scale_evidence
    assert evidence is not None
    assert evidence.derived_real_units_per_pdf_point() == 0.5

    bad_factor = copy.deepcopy(payload)
    bad_factor["assets"][0]["scale_evidence"][
        "real_units_per_pdf_point"
    ] = 0.75
    with pytest.raises(
        ValueError,
        match="does not match independently derived scale evidence",
    ):
        TakeoffDocument.model_validate(bad_factor)

    bad_geometry = copy.deepcopy(payload)
    bad_geometry["assets"][0]["scale_evidence"]["calibration"][
        "end"
    ] = {"x": 180, "y": 10}
    with pytest.raises(
        ValueError,
        match="does not match independently derived scale evidence",
    ):
        TakeoffDocument.model_validate(bad_geometry)


def test_stated_scale_ratio_independently_derives_run_quantity(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = linear_takeoff_payload(make_source_pdf(drawings))
    factor = 1 / (0.25 * 72)
    quantity = 80 * factor
    evidence = payload["assets"][0]["scale_evidence"]
    evidence["kind"] = "stated_scale"
    evidence["source_text"] = '1/4" = 1 ft'
    evidence["real_units_per_pdf_point"] = factor
    evidence.pop("calibration")
    evidence["stated_ratio"] = {
        "paper_length": 0.25,
        "paper_unit": "in",
        "real_length": 1,
        "real_unit": "ft",
    }
    payload["assets"][0]["quantity"] = quantity
    payload["by_code"][0]["quantity"] = quantity
    payload["by_area"][0]["quantity"] = quantity
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    document, _pages = validate_takeoff_artifact(
        path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )

    scale = document.assets[0].scale_evidence
    assert scale is not None
    assert scale.derived_real_units_per_pdf_point() == pytest.approx(factor)
    assert document.assets[0].quantity == pytest.approx(quantity)

    bad_factor = copy.deepcopy(payload)
    bad_factor["assets"][0]["scale_evidence"][
        "real_units_per_pdf_point"
    ] = factor * 2
    with pytest.raises(
        ValueError,
        match="does not match independently derived scale evidence",
    ):
        TakeoffDocument.model_validate(bad_factor)


@pytest.mark.parametrize(
    ("mutation", "message"),
    [
        (
            lambda payload: payload["assets"][0].update({"quantity": 41}),
            "path-length times scale",
        ),
        (
            lambda payload: payload["assets"][0]["path"].append(
                {"x": 201, "y": 80}
            ),
            "path point",
        ),
        (
            lambda payload: payload["assets"][0]["scale_evidence"].update(
                {"sheet": "E-102"}
            ),
            "same source page and sheet",
        ),
    ],
)
def test_linear_run_rejects_unverifiable_measurements(
    tmp_path: Path,
    mutation: object,
    message: str,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = linear_takeoff_payload(make_source_pdf(drawings))
    mutation(payload)
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ArtifactValidationError, match=message):
        validate_takeoff_artifact(
            path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )


def test_asset_must_map_exactly_to_source_backed_legend_entry(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    original = takeoff_payload(make_source_pdf(drawings))

    for change in (
        {"legend_entry_id": "UNKNOWN"},
        {"description": "Invented description"},
    ):
        payload = copy.deepcopy(original)
        payload["assets"][0].update(change)
        path = artifacts / "takeoff.json"
        path.write_text(json.dumps(payload), encoding="utf-8")
        with pytest.raises(
            ArtifactValidationError,
            match="does not satisfy the output schema",
        ):
            validate_takeoff_artifact(
                path,
                drawings_path=drawings,
                artifacts_dir=artifacts,
                inputs_dir=inputs,
            )


@pytest.mark.parametrize(
    "change",
    [
        {"quantity": 2, "unit": "EA"},
        {"quantity": 1, "unit": "ea"},
    ],
)
def test_count_asset_is_exactly_one_ea_placement(
    change: dict[str, object],
) -> None:
    payload = takeoff_payload("a" * 64)
    payload["assets"][0].update(change)

    with pytest.raises(
        ValueError,
        match="exactly one placement with quantity 1 and unit EA",
    ):
        TakeoffDocument.model_validate(payload)


def test_summaries_never_mix_counts_and_linear_units(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = linear_takeoff_payload(make_source_pdf(drawings))
    payload["assets"].append(
        {
            "unit_id": "TEST-CBL-COUNT-001",
            "legend_entry_id": "LEGEND-CBL",
            "measurement_kind": "count",
            "code": "CBL",
            "description": "Type C cable",
            "page": 1,
            "sheet": "E-101",
            "area_code": "L1",
            "area": "Level 1",
            "level": "1",
            "method": "symbol count",
            "confidence": "high",
            "x": 140,
            "y": 40,
            "quantity": 1,
            "unit": "EA",
        }
    )
    payload["by_code"] = [
        {
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "linear",
            "unit": "ft",
            "quantity": 41,
        }
    ]
    payload["by_area"] = [
        {
            "area_code": "L1",
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "linear",
            "unit": "ft",
            "quantity": 41,
        }
    ]
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(
        ArtifactValidationError,
        match="summary dimensions do not match assets",
    ):
        validate_takeoff_artifact(
            path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )

    payload["by_code"] = [
        {
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "count",
            "unit": "EA",
            "quantity": 1,
        },
        {
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "linear",
            "unit": "ft",
            "quantity": 40,
        },
    ]
    payload["by_area"] = [
        {
            "area_code": "L1",
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "count",
            "unit": "EA",
            "quantity": 1,
        },
        {
            "area_code": "L1",
            "legend_entry_id": "LEGEND-CBL",
            "code": "CBL",
            "description": "Type C cable",
            "measurement_kind": "linear",
            "unit": "ft",
            "quantity": 40,
        },
    ]
    path.write_text(json.dumps(payload), encoding="utf-8")

    document, _pages = validate_takeoff_artifact(
        path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )
    assert len(document.by_code) == 2
    assert {row["unit"] for row in document.by_code} == {"EA", "ft"}


def test_summaries_keep_reused_codes_separate_by_legend_definition(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    payload["legend_entries"].append(
        {
            "legend_entry_id": "LEGEND-DOOR-DOUBLE",
            "code": "DOOR",
            "description": "Double door",
            "page": 1,
            "sheet": "A-001",
            "bbox": {"x0": 20, "y0": 5, "x1": 30, "y1": 15},
        }
    )
    payload["assets"].append(
        {
            "unit_id": "TEST-DOOR-002",
            "legend_entry_id": "LEGEND-DOOR-DOUBLE",
            "measurement_kind": "count",
            "code": "DOOR",
            "description": "Double door",
            "page": 1,
            "sheet": "A-101",
            "area_code": "L1",
            "area": "Level 1",
            "level": "1",
            "method": "symbol count",
            "confidence": "high",
            "x": 100,
            "y": 40,
            "quantity": 1,
            "unit": "EA",
        }
    )
    payload["by_code"][0]["quantity"] = 2
    payload["by_area"][0]["quantity"] = 2
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(
        ArtifactValidationError,
        match="summary dimensions do not match assets",
    ):
        validate_takeoff_artifact(
            path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )

    second_dimensions = {
        "legend_entry_id": "LEGEND-DOOR-DOUBLE",
        "code": "DOOR",
        "description": "Double door",
        "measurement_kind": "count",
        "unit": "EA",
        "quantity": 1,
    }
    payload["by_code"][0]["quantity"] = 1
    payload["by_code"].append(second_dimensions)
    payload["by_area"][0]["quantity"] = 1
    payload["by_area"].append(
        {"area_code": "L1", **second_dimensions}
    )
    path.write_text(json.dumps(payload), encoding="utf-8")

    document, _pages = validate_takeoff_artifact(
        path,
        drawings_path=drawings,
        artifacts_dir=artifacts,
        inputs_dir=inputs,
    )
    assert len(document.by_code) == 2
    assert {
        row["legend_entry_id"] for row in document.by_code
    } == {"LEGEND-DOOR", "LEGEND-DOOR-DOUBLE"}


def test_takeoff_rejects_bad_summary_and_out_of_page_geometry(
    tmp_path: Path,
) -> None:
    inputs = tmp_path / "inputs"
    artifacts = tmp_path / "artifacts"
    inputs.mkdir()
    artifacts.mkdir()
    drawings = inputs / "drawings.pdf"
    payload = takeoff_payload(make_source_pdf(drawings))
    payload["assets"][0]["x"] = 500
    payload["by_code"][0]["quantity"] = 99
    path = artifacts / "takeoff.json"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(ArtifactValidationError):
        validate_takeoff_artifact(
            path,
            drawings_path=drawings,
            artifacts_dir=artifacts,
            inputs_dir=inputs,
        )


def test_workbook_rejects_corrupt_container(tmp_path: Path) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    workbook_path.write_bytes(b"PK corrupt")
    with pytest.raises(ArtifactValidationError):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_workbook_cell_limit_survives_missing_dimensions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    remove_worksheet_dimensions(workbook_path)
    monkeypatch.setattr("app.validation.MAX_WORKBOOK_CELLS", 39)

    with pytest.raises(
        ArtifactValidationError,
        match="cell inspection limit",
    ):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_workbook_row_limit_survives_missing_dimensions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    remove_worksheet_dimensions(workbook_path)
    monkeypatch.setattr("app.validation.MAX_WORKBOOK_ROWS", 1)

    with pytest.raises(
        ArtifactValidationError,
        match="worksheet bounds",
    ):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


@pytest.mark.parametrize(
    "formula",
    [
        '=WEBSERVICE("https://attacker.invalid/")',
        "=cmd|' /C calc'!A0",
    ],
)
def test_workbook_rejects_every_formula_including_dde(
    tmp_path: Path,
    formula: str,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    workbook = load_workbook(workbook_path)
    workbook["Takeoff"]["N2"] = formula
    workbook.save(workbook_path)
    workbook.close()
    with pytest.raises(ArtifactValidationError, match="not permitted"):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_workbook_rejects_single_quoted_external_relationship(
    tmp_path: Path,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    add_external_relationship(workbook_path)

    with pytest.raises(
        ArtifactValidationError,
        match="external or active relationships",
    ):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_workbook_rejects_normal_defined_name_formula(
    tmp_path: Path,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    workbook = load_workbook(workbook_path)
    workbook.defined_names.add(
        DefinedName(
            "OrdinaryCalculation",
            attr_text="SUM(Takeoff!$L$2)",
        )
    )
    workbook.save(workbook_path)
    workbook.close()

    with pytest.raises(
        ArtifactValidationError,
        match="defined names are not permitted",
    ):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_workbook_rejects_utf16_defined_name_formula(
    tmp_path: Path,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    payload = takeoff_payload("a" * 64)
    document = TakeoffDocument.model_validate(payload)
    workbook_path = artifacts / "takeoff.xlsx"
    make_workbook(workbook_path, payload)
    workbook = load_workbook(workbook_path)
    workbook.defined_names.add(
        DefinedName(
            "EncodedCalculation",
            attr_text="SUM(Takeoff!$L$2)",
        )
    )
    workbook.save(workbook_path)
    workbook.close()
    rewrite_workbook_definition_as_utf16(workbook_path)

    with pytest.raises(
        ArtifactValidationError,
        match="defined names are not permitted",
    ):
        validate_workbook_artifact(
            workbook_path, takeoff=document, artifacts_dir=artifacts
        )


def test_symlink_artifacts_are_never_read_or_registered(
    tmp_path: Path,
) -> None:
    artifacts = tmp_path / "artifacts"
    artifacts.mkdir()
    outside = tmp_path / "secret"
    outside.write_text("CODEX_API_KEY=should-not-leak", encoding="utf-8")
    link = artifacts / "takeoff.json"
    link.symlink_to(outside)
    with pytest.raises(ArtifactValidationError, match="non-symlink"):
        validate_json_artifact(link, artifacts)

    settings = Settings(
        data_dir=tmp_path / "data",
        codex_bin="codex",
        default_model="gpt-5.6-sol",
        max_upload_bytes=10_000,
        max_total_upload_bytes=10_000,
        service_api_token=None,
        max_workers=1,
        environment="test",
    )
    store = JobStore(settings.data_dir / "jobs")
    job_dir = store.create(
        JobRecord(
            id="symlinkjob",
            status=JobStatus.running,
            model=settings.default_model,
        )
    )
    artifact_link = job_dir / "artifacts" / "takeoff.json"
    artifact_link.symlink_to(outside)
    manager = PipelineManager(settings, store)
    with pytest.raises(ArtifactValidationError):
        manager._artifact("symlinkjob", artifact_link)


def test_exact_credential_material_is_rejected(tmp_path: Path) -> None:
    secret = "sk-test-not-a-real-key"
    artifact = tmp_path / "methodology.json"
    artifact.write_text(
        json.dumps({"note": f"stolen {secret}"}),
        encoding="utf-8",
    )
    with pytest.raises(ArtifactValidationError, match="credential material"):
        reject_secret_material(artifact, secret=secret)


def test_optional_pricing_metadata_is_allowed_but_bounded() -> None:
    payload = takeoff_payload("a" * 64)
    payload["assets"][0]["supplier"] = "Verified supplier"
    payload["assets"][0]["unit_price_dop"] = 1250.5
    document = TakeoffDocument.model_validate(payload)
    assert document.assets[0].supplier == "Verified supplier"

    payload["assets"][0]["supplier"] = "x" * 4_001
    with pytest.raises(ValueError, match="4000 characters"):
        TakeoffDocument.model_validate(payload)
