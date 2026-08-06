from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.models import TakeoffDocument
from app.validation import validate_workbook_artifact
from app.workbook import (
    AUDIT_HEADERS,
    SHEET_COLUMN_WIDTHS,
    SUMMARY_HEADERS,
    TABLE_NAMES,
    build_takeoff_workbook,
)


EXPECTED_VISIBLE_SHEETS = [
    "Resumen Takeoff",
    "Codigos Elementos",
    "Registro Activos",
    "Hojas de Dibujo",
    "Inventario Edificios",
    "Registros Edificios",
    "Por Tipo Edificio",
    "Por Piso",
    "Por Area",
    "Analisis Cantidades",
    "Preguntas",
    "Panel",
    "Breaker",
    "Analisis Unifilar",
    "Distancias Cableado",
    "Cronograma Equipos",
    "Cadena Unifilar",
    "Estimados Cableado",
    "Estado Entregables",
]

EXPECTED_HEADERS = {
    "Resumen Takeoff": SUMMARY_HEADERS,
    "Codigos Elementos": (
        "ID", "Numero", "Codigo de Area", "Area", "Tipo", "Piso", "Hoja",
        "Hoja de Piso", "Nombre de Plano", "Cantidad", "Descripcion", "Elemento",
        "Coordenadas",
    ),
    "Registro Activos": (
        "ID Activo", "Etiqueta Visible", "Numero", "Codigo", "Elemento",
        "Descripcion", "Hoja", "Hoja de Piso", "Nombre de Plano", "Pagina", "X90",
        "Y90", "Ancho 90", "Alto 90", "Coordenadas", "Edificio",
        "Tipo de Edificio", "Piso", "Tipo de Apartamento", "Codigo de Area", "Area",
        "Base de Area", "Distancia Etiqueta Area", "Grupo de Conteo", "Metodo Fuente",
        "Estado QA", "Notas",
    ),
    "Hojas de Dibujo": (
        "Pagina", "Numero de Hoja", "Nombre de Plano", "Clasificacion", "Piso",
        "Tipo de Apartamento", "Contable", "Notas",
    ),
    "Inventario Edificios": (
        "Tipo de Edificio", "Cantidad Observada", "Cantidad Rango Inferido",
        "Cantidad Seleccionada", "IDs Faltantes o Sospechosos", "Paginas Fuente",
    ),
    "Registros Edificios": (
        "Nombre de Edificio", "Tipo de Edificio", "Estado Fuente", "Base",
    ),
    "Por Tipo Edificio": ("Tipo", "Codigo", "Descripcion", "Cantidad"),
    "Por Piso": ("Piso", "Codigo", "Descripcion", "Cantidad"),
    "Por Area": ("Codigo de Area", "Area", "Codigo", "Descripcion", "Cantidad"),
    "Analisis Cantidades": (
        "Codigo de Area", "Area", "Tipo", "Tipo de Apartamento", "Codigo",
        "Prefijo Visible", "Descripcion", "Cantidad", "Base",
    ),
    "Preguntas": ("Pregunta", "Respuesta", "Unidad", "Base", "Nota QA"),
    "Panel": ("Indicador", "Valor"),
    "Breaker": ("Seccion", "Dato", "Panel", "Elemento"),
    "Analisis Unifilar": (
        "Tipo de Edificio", "Cantidad Observada", "Cantidad Rango Inferido",
        "Cantidad Seleccionada", "IDs Faltantes o Sospechosos", "Paginas Fuente", "Base",
    ),
    "Distancias Cableado": (
        "Segmento", "Desde Nodo", "Hasta Nodo", "Distancia", "Unidad", "Estado", "Base",
        "Hoja", "Notas",
    ),
    "Cronograma Equipos": (
        "ID Equipo", "Tipo de Equipo", "Descripcion", "Capacidad/Rating",
        "Cantidad/Capacidad", "Edificio", "Piso", "Hoja", "Texto Fuente", "Notas",
    ),
    "Cadena Unifilar": (
        "ID Cadena", "Desde Nodo", "Hasta Nodo", "Tipo de Nodo", "Referencia Equipo",
        "Capacidad/Rating", "Hoja", "Coordenadas", "Notas",
    ),
    "Estimados Cableado": (
        "ID Cable", "Tipo de Segmento", "Desde Nodo", "Hasta Nodo", "Longitud Estimada",
        "Unidad", "Tipo de Cable", "Base", "Hoja", "Notas",
    ),
    "Estado Entregables": ("Entregable", "Estado", "Archivo", "Notas"),
}


def sample_takeoff() -> TakeoffDocument:
    return TakeoffDocument.model_validate(
        {
            "source": {
                "pdf": "sample-electrical.pdf",
                "sha256": "a" * 64,
                "page_count": 2,
            },
            "legend_entries": [
                {
                    "legend_entry_id": "LEGEND-TC",
                    "code": "TC",
                    "description": "Duplex receptacle",
                    "page": 1,
                    "sheet": "E-001",
                    "bbox": {"x0": 1, "y0": 1, "x1": 9, "y1": 9},
                },
                {
                    "legend_entry_id": "LEGEND-CBL",
                    "code": "CBL",
                    "description": "Type C cable",
                    "page": 2,
                    "sheet": "E-002",
                    "bbox": {"x0": 1, "y0": 1, "x1": 9, "y1": 9},
                },
            ],
            "assets": [
                {
                    "unit_id": "TC-001",
                    "legend_entry_id": "LEGEND-TC",
                    "measurement_kind": "count",
                    "code": "TC",
                    "description": "Duplex receptacle",
                    "page": 1,
                    "sheet": "E-101",
                    "area_code": "L1-A",
                    "area": "Level 1 - Area A",
                    "level": "1",
                    "method": "visual count",
                    "confidence": "high",
                    "bbox": {"x0": 10, "y0": 10, "x1": 20, "y1": 20},
                    "visible_label": "TC",
                    "quantity": 1,
                    "unit": "EA",
                },
                {
                    "unit_id": "CBL-001",
                    "legend_entry_id": "LEGEND-CBL",
                    "measurement_kind": "linear",
                    "code": "CBL",
                    "description": "Type C cable",
                    "page": 2,
                    "sheet": "E-102",
                    "area_code": "L2-A",
                    "area": "Level 2 - Area A",
                    "level": "2",
                    "method": "scaled centerline",
                    "confidence": "medium",
                    "path": [{"x": 20, "y": 30}, {"x": 30, "y": 30}],
                    "scale_evidence": {
                        "kind": "calibrated_dimension",
                        "page": 2,
                        "sheet": "E-102",
                        "bbox": {"x0": 40, "y0": 40, "x1": 50, "y1": 50},
                        "source_text": "5 ft calibration",
                        "unit": "ft",
                        "real_units_per_pdf_point": 0.5,
                        "calibration": {
                            "start": {"x": 40, "y": 45},
                            "end": {"x": 50, "y": 45},
                            "known_length": 5,
                            "unit": "ft",
                        },
                    },
                    "quantity": 5,
                    "unit": "ft",
                },
            ],
            "unresolved_symbols": [],
            "by_code": [],
            "by_area": [],
            "limitations": ["Confirm final device subtypes against the legend."],
        }
    )


def test_builds_static_ortega_format_and_hidden_audit(tmp_path: Path) -> None:
    takeoff = sample_takeoff()
    output = tmp_path / "takeoff.xlsx"

    build_takeoff_workbook(takeoff, output)
    validate_workbook_artifact(output, takeoff=takeoff, artifacts_dir=tmp_path)

    workbook = load_workbook(output, data_only=False, keep_links=False)
    try:
        assert workbook.sheetnames == [*EXPECTED_VISIBLE_SHEETS, "Takeoff"]
        assert all(
            workbook[name].sheet_state == "visible"
            for name in EXPECTED_VISIBLE_SHEETS
        )
        assert workbook["Takeoff"].sheet_state == "hidden"
        for sheet_name in EXPECTED_VISIBLE_SHEETS:
            worksheet = workbook[sheet_name]
            assert tuple(cell.value for cell in worksheet[1]) == EXPECTED_HEADERS[sheet_name]
            assert worksheet.freeze_panes is None
            assert worksheet.sheet_view.showGridLines is False
            assert worksheet.row_dimensions[1].height == 32
            assert len(worksheet.tables) == 1
            table = next(iter(worksheet.tables.values()))
            assert table.displayName == TABLE_NAMES[sheet_name]
            assert table.tableStyleInfo.name == "TableStyleMedium2"
            for column_number, width in enumerate(
                SHEET_COLUMN_WIDTHS[sheet_name], start=1
            ):
                assert worksheet.column_dimensions[
                    get_column_letter(column_number)
                ].width == width
            header = worksheet["A1"]
            assert header.fill.fgColor.rgb == "006FA8DC"
            assert header.font.name == "Carlito"
            assert header.font.sz == 11
            assert header.font.bold is True
            assert header.alignment.horizontal == "center"
            assert header.alignment.wrap_text is True
        assert tuple(cell.value for cell in workbook["Takeoff"][1]) == AUDIT_HEADERS
        assert workbook["Takeoff"].max_row == 3
        assert workbook["Takeoff"]["A2"].value == "TC-001"
        assert workbook["Takeoff"]["A3"].value == "CBL-001"
        assert workbook["Takeoff"]["O3"].value == 10
        assert workbook["Takeoff"]["T3"].value == 0.5

        summary_quantities = {
            row[4].value: row[5].value
            for row in workbook["Resumen Takeoff"].iter_rows(min_row=2)
        }
        assert summary_quantities == {
            "CBL — Type C cable": 5,
            "TC — Duplex receptacle": 1,
        }
        assert workbook["Distancias Cableado"]["A2"].value == "CBL-001"
        assert workbook["Estimados Cableado"]["E2"].value == 5
        assert workbook["Por Piso"].max_row == 3
        assert workbook["Por Piso"]["C2"].value == "Duplex receptacle [EA]"
        assert workbook["Por Piso"]["C3"].value == "Type C cable [ft]"
        panel_values = {
            row[0].value: row[1].value
            for row in workbook["Panel"].iter_rows(min_row=2)
        }
        assert "Cantidad total" not in panel_values
        assert panel_values["Conteo total (EA)"] == 1
        assert panel_values["Longitud total (ft)"] == 5
        assert panel_values["TC (EA)"] == 1
        assert panel_values["CBL (ft)"] == 5
        assert workbook["Analisis Cantidades"]["C2"].value == "Conteo (EA)"
        assert workbook["Analisis Cantidades"]["C3"].value == "Lineal (ft)"
        assert workbook["Resumen Takeoff"]["C2"].value == "LEGEND-CBL"
        assert workbook["Registro Activos"]["K2"].value == 18.75
        assert workbook["Registro Activos"]["M2"].value == 12.5
        assert workbook["Codigos Elementos"]["M2"].value == (
            "x0=12.5; y0=12.5; x1=25.0; y1=25.0"
        )
        assert workbook["Preguntas"]["A2"].fill.fgColor.rgb == "00FFF2CC"
        assert workbook["Preguntas"]["A2"].alignment.wrap_text is True
        assert workbook["Panel"]["A2"].alignment.wrap_text is True
        assert workbook["Resumen Takeoff"]["A2"].fill.fill_type is None
        assert len(workbook.defined_names) == 0
        assert workbook._external_links == []

        formulas = [
            cell.coordinate
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
            or isinstance(cell.value, str)
            and cell.value.startswith("=")
        ]
        assert formulas == []
    finally:
        workbook.close()
